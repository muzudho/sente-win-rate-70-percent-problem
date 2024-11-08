#
# python step_oa44o1o0_manual_report_vrs.py
#
#   VRS表を CSV形式から XLSX形式へ変換します
#

import traceback
import os
import random
import math
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

from library import HEAD, TAIL, toss_a_coin
from library.file_paths import VictoryRateSummaryFilePaths


########################################
# コマンドから実行時
########################################


if __name__ == '__main__':
    """コマンドから実行時"""

    try:

        # ファイルパス取得
        # ---------------
        csv_file_path = VictoryRateSummaryFilePaths.as_csv()                # VRS表 CSV
        wb_file_path = VictoryRateSummaryFilePaths.as_workbook_on_reports() # ワークブック

        # CSV読取
        # -------
        if os.path.isfile(csv_file_path):   # VRS表(CSV)。既存時
            df = pd.read_csv(
                    csv_file_path,
                    encoding="utf-8")
        
            # 列順の変更
            df = df[[
                'p',                        # 先手勝率
                'failure_rate',             # 引分け率
                'turn_system_name',         # 手番の決め方
                'h_step',                   # 表点
                't_step',                   # ｳﾗ点
                'span',                     # 優勝点
                'h_time',                   # 必要表回数
                't_time',                   # 必要ｳﾗ回数
                'shortest_coins',           # 最短対局数
                'upper_limit_coins',        # 対局数上限
                'a_victory_rate_by_trio',
                'b_victory_rate_by_trio',
                'no_victory_rate',
                'a_victory_rate_by_duet',
                'b_victory_rate_by_duet',
                'unfair_point']]

            # ソート
            df.sort_values(
                    by=['p','failure_rate','turn_system_name'],
                    ascending=[
                        True,
                        True,
                        False], # 固定→交互の順
                    inplace=True)

        else:
            raise ValueError(f"file not found {csv_file_path=}")


        # ワークブックの新規作成
        wb = xl.Workbook()

        # 既定シート名変更
        wb['Sheet'].title = 'Summary'

        # シート取得
        ws = wb['Summary']

        # 列幅設定  数値通りにはいかないようだ
        column_width_list = [
            ('A', 8.64),    # ［先手勝率］
            ('B', 6.64),    # 引分率
            ('C', 12.73),   # 手番の決め方
            ('D', 4.64),    # 表点
            ('E', 5.00),    # ｳﾗ点
            ('F', 6.64),    # 優勝点
            ('G', 10.73),   # 必要表回数
            ('H', 11.18),   # 必要ｳﾗ回数
            ('I', 10.73),   # 最短対局数
            ('J', 10.73),   # 対局数上限
            ('K', 5.0),     # ［Ａさんの優勝率（優勝なし率込）］
            ('L', 5.0),     # ［Ｂさんの優勝率（優勝なし率込）］
            ('M', 16.55),   # ［優勝なし率］
            ('N', 13.91),   # ［Ａさんの優勝率］
            ('O', 14.00),   # ［Ｂさんの優勝率］
            ('P', 11.91),   # ［不均等度］
        ]
        for pair in column_width_list:
            ws.column_dimensions[pair[0]].width = pair[1]

        # 列名の変更
        new_column_name_dict = {
            'p':'先手勝率',
            'failure_rate':'引分率',
            'turn_system_name':'手番の決め方',
            'h_step':'表点',
            't_step':'ｳﾗ点',
            'span':'優勝点',
            'h_time':'必要表回数',
            't_time':'必要ｳﾗ回数',
            'shortest_coins':'最短対局数',
            'upper_limit_coins':'対局数上限',
            'a_victory_rate_by_trio':'Ａさんの優勝率（優勝なし率込）',
            'b_victory_rate_by_trio':'Ｂさんの優勝率（優勝なし率込）',
            'no_victory_rate':'優勝なし率',
            'a_victory_rate_by_duet':'Ａさんの優勝率',
            'b_victory_rate_by_duet':'Ｂさんの優勝率',
            'unfair_point':'不均等度'}

        # 寄せ
        alignment_list = [
            Alignment(horizontal='left'),   # 先手勝率
            Alignment(horizontal='left'),   # 引分率
            Alignment(horizontal='left'),   # 手番の決め方
            Alignment(horizontal='right'),  # 表点
            Alignment(horizontal='right'),  # ｳﾗ点
            Alignment(horizontal='right'),  # 優勝点
            Alignment(horizontal='right'),  # 必要表回数
            Alignment(horizontal='right'),  # 必要ｳﾗ回数
            Alignment(horizontal='right'),  # 最短対局数
            Alignment(horizontal='right'),  # 対局数上限
            Alignment(horizontal='left'),   # Ａさんの優勝率（優勝なし率込）
            Alignment(horizontal='left'),   # Ｂさんの優勝率（優勝なし率込）
            Alignment(horizontal='left'),   # 優勝なし率
            Alignment(horizontal='left'),   # Ａさんの優勝率
            Alignment(horizontal='left'),   # Ｂさんの優勝率
            Alignment(horizontal='left')]   # 不均等度
        

        # ヘッダー文字色・背景色
        header_font = Font(color='EEFFEE')
        header_background_fill = PatternFill(patternType='solid', fgColor='336633')

        # ヘッダー出力
        for column_th, column_name in enumerate(df.columns.values, 1):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}1']
            cell.value = new_column_name_dict[column_name]
            cell.font = header_font
            cell.fill = header_background_fill
            cell.alignment = alignment_list[column_th - 1]


        # データ部
        # -------
        column_letter_and_column_name = [
            ('A', 'p'),                         # 先手勝率
            ('B', 'failure_rate'),              # 引分け率
            ('C', 'turn_system_name'),          # 手番の決め方
            ('D', 'h_step'),                    # 表点
            ('E', 't_step'),                    # ｳﾗ点
            ('F', 'span'),                      # 優勝点
            ('G', 'h_time'),                    # 必要表回数
            ('H', 't_time'),                    # 必要ｳﾗ回数
            ('I', 'shortest_coins'),            # 最短対局数
            ('J', 'upper_limit_coins'),         # 対局数上限
            ('K', 'a_victory_rate_by_trio'),
            ('L', 'b_victory_rate_by_trio'),
            ('M', 'no_victory_rate'),
            ('N', 'a_victory_rate_by_duet'),
            ('O', 'b_victory_rate_by_duet'),
            ('P', 'unfair_point'),
        ]

        turn_system_name_to_jp = {
            'alternating':'交互',
            'frozen':'固定',
        }

        for sorted_row_no, (row_no, row) in enumerate(df.iterrows()):
            #print(f"{row_th=} {type(row)=}")
            row_th = sorted_row_no + 2

            for column_no in range(0, len(row)):
                cell = ws[f'{column_letter_and_column_name[column_no][0]}{row_th}']
                cell.value = row[column_letter_and_column_name[column_no][1]]
                cell.alignment = alignment_list[column_no]

                # ［手番の決め方］
                if column_no == 2:
                    cell.value = turn_system_name_to_jp[cell.value]


        # ウィンドウ枠の固定
        ws.freeze_panes = 'A2'


        # データ部背景色
        # -------------
        light_red_fill = PatternFill(patternType='solid', fgColor='FFEEEE')
        light_yellow_fill = PatternFill(patternType='solid', fgColor='FFFFEE')


        # データ部罫線出力
        # ---------------
        black_side = Side(style='thick', color='333333')
        left_border = Border(left=black_side)
        right_border = Border(right=black_side)
        top_border = Border(top=black_side)

        pre_p = None

        for row_th in range(2, ws.max_row + 1):

            # ［先手勝率］
            cell = ws[f'A{row_th}']
            # ［先手勝率］の変わり目
            if pre_p is not None and pre_p != cell.value:
                cell.border = top_border

            pre_p = cell.value


            # ［表点］
            cell = ws[f'D{row_th}']
            cell.border = left_border
            cell.fill = light_red_fill

            # ［ｳﾗ点］
            cell = ws[f'E{row_th}']
            cell.fill = light_red_fill

            # ［優勝点］
            cell = ws[f'F{row_th}']
            cell.border = right_border
            cell.fill = light_red_fill

            # ［Ａさんの優勝率（優勝なし率込）］
            cell = ws[f'K{row_th}']
            cell.border = left_border

            # ［Ａさんの優勝率］
            cell = ws[f'N{row_th}']
            cell.border = left_border
            cell.fill = light_yellow_fill

            # ［Ｂさんの優勝率］
            cell = ws[f'O{row_th}']
            cell.border = right_border
            cell.fill = light_yellow_fill


        # ワークブックの保存
        wb.save(wb_file_path)


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
