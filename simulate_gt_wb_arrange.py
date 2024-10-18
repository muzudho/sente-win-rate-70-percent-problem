#
# 分析
# python simulate_gt_wb_arrange.py
#
#   GTWB をアレンジします
#

import traceback
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

from library import HEAD, TAIL, Specification, SeriesRule
from library.database import GameTreeNode, GameTreeRecord, GameTreeTable
from library.workbooks import GameTreeWorkbookWrapper
from library.views import PromptCatalog
from library.game_tree_view import GameTreeView


class TreeDrawer():
    """エクセルで罫線などを駆使して、樹形図を描画します"""


    def __init__(self, gt_table, gt_wb_wrapper):
        self._gt_table = gt_table
        self._gt_wb_wrapper = gt_wb_wrapper
        self._prev_gt_record = GameTreeRecord.new_empty()
        self._curr_gt_record = GameTreeRecord.new_empty()
        self._next_gt_record = GameTreeRecord.new_empty()


    def forward_cursor(self, next_gt_record):
        """送り出し"""
        self._prev_gt_record = self._curr_gt_record
        self._curr_gt_record = self._next_gt_record
        self._next_gt_record = next_gt_record


    def on_header(self):

        # 変数名の短縮
        ws = self._gt_wb_wrapper.worksheet


        # 列の幅設定
        # width はだいたい 'ＭＳ Ｐゴシック' サイズ11 の半角英文字の個数

        # TODO C列には確率を入れたい
        # TODO D列は空列にしたい
        # TODO E列の上の方の行には 1 を入れたい

        ws.column_dimensions['A'].width = 4     # no
        ws.column_dimensions['B'].width = 20    # result
        ws.column_dimensions['C'].width = 14    # rate
        ws.column_dimensions['D'].width = 14    # empty column
        ws.column_dimensions['E'].width = 14    # root node
        ws.column_dimensions['F'].width = 2    # 1
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 2    # 2
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 2    # 3
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 10
        ws.column_dimensions['O'].width = 2    # 4
        ws.column_dimensions['P'].width = 14
        ws.column_dimensions['Q'].width = 10
        ws.column_dimensions['R'].width = 2    # 5
        ws.column_dimensions['S'].width = 14
        ws.column_dimensions['T'].width = 10
        ws.column_dimensions['U'].width = 2    # 6
        ws.column_dimensions['V'].width = 14
        ws.column_dimensions['W'].width = 10


        # 行の高さ設定
        # height の単位はポイント。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
        ws.row_dimensions[1].height = 13
        ws.row_dimensions[2].height = 13


        # 第１行
        # ------
        # ヘッダー行にする
        row_th = 1

        # そのままコピーできない
        # # ２列目～
        # for column_number, column_name in enumerate(self._gt_table.df.columns.values, 2):
        #     ws[f'{xl.utils.get_column_letter(column_number)}{row_th}'] = column_name
        ws[f'{xl.utils.get_column_letter(1)}{row_th}'] = 'No'
        ws[f'{xl.utils.get_column_letter(2)}{row_th}'] = '結果'
        ws[f'{xl.utils.get_column_letter(3)}{row_th}'] = '実現確率'
        # 4 は空列

        ws[f'{xl.utils.get_column_letter(5)}{row_th}'] = '開始前'

        # 6 は分岐線
        # 7 はedge
        ws[f'{xl.utils.get_column_letter(8)}{row_th}'] = '1局後'   # node
        ws[f'{xl.utils.get_column_letter(11)}{row_th}'] = '2局後'
        ws[f'{xl.utils.get_column_letter(14)}{row_th}'] = '3局後'
        ws[f'{xl.utils.get_column_letter(17)}{row_th}'] = '4局後'
        ws[f'{xl.utils.get_column_letter(20)}{row_th}'] = '5局後'
        ws[f'{xl.utils.get_column_letter(23)}{row_th}'] = '6局後'

        # 第２行
        # ------
        # 空行にする
        row_th = 2


    def on_each_gt_record(self, next_row_number, next_gt_record):
        """先読みで最初の１回を空振りさせるので、２行目から本処理です"""

        # 事前送り出し
        self.forward_cursor(next_gt_record=next_gt_record)

        curr_row_number = next_row_number - 1
        curr_row_th = curr_row_number + 1

        if self._curr_gt_record.no is None:
            print(f"[{datetime.datetime.now()}] {curr_row_th}行目 現在レコードのnoがナンだから無視（先読みのため、初回は空回し）")
            pass


        else:
            # 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
            node_bgcolor = PatternFill(patternType='solid', fgColor='FFFFCC')

            # 罫線
            #
            #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
            #
            side = Side(style='thick', color='000000')
            # デバッグ用に色を付けておく
            red_side = Side(style='thick', color='660000')      # FF0000
            orange_side = Side(style='thick', color='663300')   # FFCC00
            green_side = Side(style='thick', color='006600')    # 00FF00
            blue_side = Side(style='thick', color='000066')     # 0000FF
            # 黄色は白字の上で見にくいのでやめとく
            cyan_side = Side(style='thick', color='006666')     # 00FFFF
            magenta_side = Side(style='thick', color='660066')  # FF00FF
            # 親への接続は赤
            border_to_parent = Border(bottom=red_side)
            # 子への水平接続はオレンジ
            under_border_to_child_horizontal = Border(bottom=orange_side)
            # 子へのダウン接続はブルー
            under_border_to_child_down = Border(bottom=blue_side)
            leftside_border_to_child_down = Border(left=blue_side)
            # 子へのＴ字接続はシアン
            l_letter_border_to_child_t_letter = Border(left=cyan_side, bottom=cyan_side)
            leftside_border_to_child_t_letter = Border(left=cyan_side)
            # 子へのアップ接続はグリーン
            l_letter_border_to_child_up = Border(left=green_side, bottom=green_side)
            # 垂直接続はマゼンタ
            leftside_border_to_vertical = Border(left=magenta_side)

            upside_node_border = Border(top=side, left=side, right=side)
            downside_node_border = Border(bottom=side, left=side, right=side)


            # 変数名短縮
            ws = self._gt_wb_wrapper.worksheet


            # ３行目～６行目
            # -------------
            # データは３行目から、１かたまり３行を使って描画する
            row1_th = curr_row_number * 3 + 3
            row2_th = curr_row_number * 3 + 3 + 1
            row3_th = curr_row_number * 3 + 3 + 2
            three_row_numbers = [row1_th, row2_th, row3_th]

            # 行の高さ設定
            # height の単位はポイント。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
            ws.row_dimensions[row1_th].height = 13
            ws.row_dimensions[row2_th].height = 13
            ws.row_dimensions[row3_th].height = 6

            ws[f'A{row1_th}'].value = self._curr_gt_record.no
            ws[f'B{row1_th}'].value = self._curr_gt_record.result


            # TODO C列には確率を入れたい。あとで入れる
            # TODO D列は空列にしたい
            # TODO E列の上の方の行には 1 を入れたい


            def draw_node(round_th, three_column_names, three_row_numbers):
                """
                Parameters
                ----------
                round_th : int
                    第何局後
                
                Return
                ------
                nd : GameTreeNode
                    対象ノード
                """

                round_no = round_th - 1
                prerow_nd = self._prev_gt_record.node_at(round_no=round_no)
                nd = self._curr_gt_record.node_at(round_no=round_no)

                if nd is None:
                    print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  nd がナンのノードは無視")
                    return nd

                elif pd.isnull(nd.face):
                    print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  nd.face が NaN のノードは無視")
                    return nd

                elif pd.isnull(nd.rate):
                    print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  nd.rate が NaN のノードは無視")
                    return nd

                elif prerow_nd is None:
                    # 前行が無ければ描画
                    pass


                # 以下、描画
                if curr_row_th != self._curr_gt_record.no:
                    raise ValueError(f"行番号がずれている {curr_row_th=}  {self._curr_gt_record.no=}")
                print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後 ノード描画...")


                def edge_text(node):
                    if node.face == 'h':
                        face = '表'
                    elif node.face == 't':
                        face = '裏'
                    elif node.face == 'f':
                        face = '失敗'
                    else:
                        raise ValueError(f"{node.face=}")
                    
                    if node.winner == 'A':
                        winner = '(Ａさん'
                    elif node.winner == 'B':
                        winner = '(Ｂさん'
                    elif node.winner == 'N':
                        winner = ''
                    else:
                        raise ValueError(f"{node.winner=}")

                    if node.pts != -1:
                        pts = f"{node.pts:.0f}点)" # FIXME 小数部を消してる。これで誤差で丸めを間違えるケースはあるか？
                    else:
                        pts = ''

                    return f"{face}{winner}{pts}"

                cn1 = three_column_names[0]
                cn2 = three_column_names[1]
                cn3 = three_column_names[2]
                row1_th = three_row_numbers[0]
                row2_th = three_row_numbers[1]
                row3_th = three_row_numbers[2]


                if prerow_nd is not None and nd.rate == prerow_nd.rate:

                    # 垂直線
                    #
                    #   |    leftside_border
                    # ..+..  
                    #   |    leftside_border
                    #   |    leftside_border
                    #
                    if GameTreeView.is_same_as_avobe(
                            curr_gt_record=self._curr_gt_record,
                            prev_gt_record=self._prev_gt_record,
                            round_th=round_th):
                        print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  垂直線")
                        
                        ws[f'{cn2}{row1_th}'].border = leftside_border_to_vertical
                        ws[f'{cn2}{row2_th}'].border = leftside_border_to_vertical
                        ws[f'{cn2}{row3_th}'].border = leftside_border_to_vertical
                    
                    else:
                        print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  空欄")
                        pass

                    return nd


                # １列目：親ノードから伸びてきた枝
                #
                #   .
                # --...
                #   .
                #
                # 前ラウンドにノードがあれば、接続線を引く
                #
                if GameTreeView.can_connect_to_parent(
                        curr_gt_record=self._curr_gt_record,
                        prev_gt_record=self._prev_gt_record,
                        round_th=round_th):
                    ws[f'{cn1}{row1_th}'].border = border_to_parent
                

                # ２列目：分岐したエッジ
                ws[f'{cn2}{row1_th}'].value = edge_text(node=nd)


                # 子ノードへの接続は４種類の線がある
                #
                # (1) Horizontal
                #   .    under_border
                # ...__  
                #   .    None
                #   .    None
                #
                # (2) Down
                #   .    under_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (3) TLetter
                #   |    l_letter_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (4) Up
                #   |    l_letter_border
                # ..+__  
                #   .    None
                #   .    None
                #
                kind = GameTreeView.get_kind_connect_to_child(
                        prev_gt_record=self._prev_gt_record,
                        curr_gt_record=self._curr_gt_record,
                        next_gt_record=self._next_gt_record,
                        round_th=round_th)

                if kind == 'Horizontal':
                    ws[f'{cn2}{row1_th}'].border = under_border_to_child_horizontal
                    print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  水平線")
                
                elif kind == 'Down':
                    ws[f'{cn2}{row1_th}'].border = under_border_to_child_down
                    ws[f'{cn2}{row2_th}'].border = leftside_border_to_child_down
                    ws[f'{cn2}{row3_th}'].border = leftside_border_to_child_down
                    print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  ダウン線")

                elif kind == 'TLetter':
                    ws[f'{cn2}{row1_th}'].border = l_letter_border_to_child_t_letter
                    ws[f'{cn2}{row2_th}'].border = leftside_border_to_child_t_letter
                    ws[f'{cn2}{row3_th}'].border = leftside_border_to_child_t_letter
                    print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  Ｔ字線")

                elif kind == 'Up':
                    ws[f'{cn2}{row1_th}'].border = l_letter_border_to_child_up
                    print(f"[{datetime.datetime.now()}] {curr_row_th}行目 {round_th}局後  アップ線")
                
                else:
                    raise ValueError(f"{kind=}")

                # ３列目：箱
                ws[f'{cn3}{row1_th}'].value = nd.rate
                ws[f'{cn3}{row1_th}'].fill = node_bgcolor
                ws[f'{cn3}{row1_th}'].border = upside_node_border
                ws[f'{cn3}{row2_th}'].fill = node_bgcolor
                ws[f'{cn3}{row2_th}'].border = downside_node_border

                return nd


            # 根ノード
            # -------
            if curr_row_number == 0:
                ws[f'E{row1_th}'].value = 1
                ws[f'E{row1_th}'].fill = node_bgcolor
                ws[f'E{row1_th}'].border = upside_node_border
                ws[f'E{row2_th}'].fill = node_bgcolor
                ws[f'E{row2_th}'].border = downside_node_border


            # それ以外のノード
            # ---------------

            # 実現確率
            rate = None


            # 1局後
            # -----
            nd = draw_node(
                    round_th=1,
                    three_column_names=['F', 'G', 'H'],
                    three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 2局後
            # -----
            nd = draw_node(
                    round_th=2,
                    three_column_names=['I', 'J', 'K'],
                    three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 3局後
            # -----
            nd = draw_node(
                    round_th=3,
                    three_column_names=['L', 'M', 'N'],
                    three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 4局後
            # -----
            nd = draw_node(
                    round_th=4,
                    three_column_names=['O', 'P', 'Q'],
                    three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 5局後
            # -----
            nd = draw_node(
                    round_th=5,
                    three_column_names=['R', 'S', 'T'],
                    three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 6局後
            # -----
            nd = draw_node(
                    round_th=6,
                    three_column_names=['U', 'V', 'W'],
                    three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 実現確率
            # --------
            ws[f'C{row1_th}'].value = rate


class TreeEraser():
    """要らない罫線を消す"""


    def __init__(self, gt_wb_wrapper):
        self._gt_wb_wrapper = gt_wb_wrapper


    def erase_unnecessary_border_by_column(self, column_alphabet):
        """不要な境界線を消す"""

        # 変数名の短縮
        ws = self._gt_wb_wrapper.worksheet

        row_th_of_last_underline = -1


        # 行番号は 4 から
        row_th = 4
        while row_th <= ws.max_row: # 最終行まで全部見る

            while True:

                # 罫線を確認
                #
                #   .
                # ..+--  下向きの罫線が最後に出た箇所を調べる
                #   |
                #
                border = ws[f'{column_alphabet}{row_th}'].border
                if border is not None:
                    #print(f"[{datetime.datetime.now()}] 消しゴム {row_th=} 境界線有り {border=}")

                    there_no_border = True

                    if border.left is not None:
                        #print(f"[{datetime.datetime.now()}] 消しゴム {row_th=} {border.left.style=}")
                        if border.left.style == 'thick':
                            there_no_border = False
                            #print(f"[{datetime.datetime.now()}] 消しゴム {row_th=} 左側に罫線")

                    if border.bottom is not None:
                        #print(f"[{datetime.datetime.now()}] 消しゴム {row_th=} {border.bottom.style=}")
                        if border.bottom.style == 'thick':
                            there_no_border = False
                            row_th_of_last_underline = row_th
                            print(f"[{datetime.datetime.now()}] 消しゴム {row_th=} アンダーライン")

                    # 境界線が無かったらループを抜ける
                    if there_no_border:
                        print(f"[{datetime.datetime.now()}] 消しゴム {row_th=} ループ抜ける {ws.max_row=}")
                        break

                row_th += 1

            print(f"[{datetime.datetime.now()}] 消しゴムを掛けたい行の番号 {row_th_of_last_underline+1}～{row_th-1}")
            # 消しゴムを掛ける
            if row_th_of_last_underline != -1:
                for temp_row_th in range(row_th_of_last_underline+1, row_th):
                    ws[f'{column_alphabet}{temp_row_th}'].border = None

            # 次行から続行
            row_th += 1


    def execute(self):

        # TODO 可変長に対応したい
        # G列の左側の垂直線を見ていく
        self.erase_unnecessary_border_by_column(column_alphabet='G')
        self.erase_unnecessary_border_by_column(column_alphabet='J')
        self.erase_unnecessary_border_by_column(column_alphabet='M')
        self.erase_unnecessary_border_by_column(column_alphabet='P')


########################################
# コマンドから実行時
########################################
if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # ［先後の決め方］を尋ねます
        specified_turn_system_id = PromptCatalog.which_method_do_you_use_to_determine_sente_and_gote()


        # ［将棋の引分け率］を尋ねます
        specified_failure_rate = PromptCatalog.what_is_the_failure_rate()


        # ［将棋の先手勝率］を尋ねます
        specified_p = PromptCatalog.what_is_the_probability_of_flipping_a_coin_and_getting_heads()


        # ［目標の点数］を尋ねます
        specified_span = PromptCatalog.how_many_goal_win_points()


        # ［後手で勝ったときの勝ち点］を尋ねます
        specified_t_step = PromptCatalog.how_many_win_points_of_tail_of_coin()


        # ［先手で勝ったときの勝ち点］を尋ねます
        specified_h_step = PromptCatalog.how_many_win_points_of_head_of_coin()


        # ［仕様］
        spec = Specification(
                turn_system_id=specified_turn_system_id,
                failure_rate=specified_failure_rate,
                p=specified_p)

        # FIXME 便宜的に［試行シリーズ数］は 1 固定
        specified_trial_series = 1

        # ［シリーズ・ルール］。任意に指定します
        specified_series_rule = SeriesRule.make_series_rule_base(
                spec=spec,
                span=specified_span,
                t_step=specified_t_step,
                h_step=specified_h_step)


        # GTWB ファイル作成
        # オリジナルはインターフェースに癖があるので、ラッパーを作成してそれを使う
        gt_wb_wrapper = GameTreeWorkbookWrapper.instantiate(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

        # ワークブックを開く（既存なら削除してから新規作成）
        gt_wb = gt_wb_wrapper.open_workbook(remove_workbook_if_it_exists=True)

        # GameTree シートを作成
        gt_wb_wrapper.create_sheet('GameTree', shall_overwrite=False)

        # 既存の Sheet シートを削除
        gt_wb_wrapper.remove_sheet('Sheet')


        # TODO 下につながらない垂直線（兄弟の末っ子から下に垂れる垂直線）を描画しないために、印を付けたい
        # NOTE プリフェッチは難しい。エクセルの罫線をスキャンした方が楽か


        # GTテーブル
        gt_table, gt_file_read_result = GameTreeTable.from_csv(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD),
                new_if_it_no_exists=False)

        if gt_file_read_result.is_file_not_found:
            raise ValueError(f"GTファイルが見つかりません {gt_file_read_result.file_path=}")


        print(f"""\
gt_table:
{gt_table}""")


        tree_drawer = TreeDrawer(gt_table=gt_table, gt_wb_wrapper=gt_wb_wrapper)

        # GTWB の Sheet シートへのヘッダー書出し
        tree_drawer.on_header()

        # GTWB の Sheet シートへの各行書出し
        gt_table.for_each(on_each=tree_drawer.on_each_gt_record)

        # 最終行の実行
        tree_drawer.on_each_gt_record(next_row_number=len(gt_table.df), next_gt_record=GameTreeRecord.new_empty())


        # 要らない罫線を消す
        tree_eraser = TreeEraser(gt_wb_wrapper=gt_wb_wrapper)
        tree_eraser.execute()


        # GTWB ファイルの保存
        gt_wb_wrapper.save()


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
