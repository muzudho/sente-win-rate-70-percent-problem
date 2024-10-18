#
# TODO 作りかけ
#
# python simulate_gt_wb_summary.py
#
#   * GTWB の Tree シートの葉要素を一覧 Leafs シート
#   * Aさんの勝率など、科目別に集計 Sum シート
#

import traceback
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

from library import HEAD, TAIL, Specification, SeriesRule
from library.file_paths import GameTreeWorkbookFilePaths, GameTreeFilePaths
from library.database import GameTreeNode, GameTreeRecord, GameTreeTable
from library.workbooks import GameTreeWorkbookWrapper
from library.views import PromptCatalog
from library.game_tree_view import GameTreeView


########################################
# コマンドから実行時
########################################
if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # ［先後の決め方］を尋ねます
        specified_turn_system_id = PromptCatalog.which_method_do_you_use_to_determine_sente_and_gote()


        # ［将棋の引分け率］を尋ねます
        specified_failure_rate = PromptCatalog.what_is_the_failure_rate()


        # ［将棋の先手勝率］を尋ねます
        specified_p = PromptCatalog.what_is_the_probability_of_flipping_a_coin_and_getting_heads()


        # ［目標の点数］を尋ねます
        specified_span = PromptCatalog.how_many_goal_win_points()


        # ［後手で勝ったときの勝ち点］を尋ねます
        specified_t_step = PromptCatalog.how_many_win_points_of_tail_of_coin()


        # ［先手で勝ったときの勝ち点］を尋ねます
        specified_h_step = PromptCatalog.how_many_win_points_of_head_of_coin()


        # ［仕様］
        spec = Specification(
                turn_system_id=specified_turn_system_id,
                failure_rate=specified_failure_rate,
                p=specified_p)

        # FIXME 便宜的に［試行シリーズ数］は 1 固定
        specified_trial_series = 1

        # ［シリーズ・ルール］。任意に指定します
        specified_series_rule = SeriesRule.make_series_rule_base(
                spec=spec,
                span=specified_span,
                t_step=specified_t_step,
                h_step=specified_h_step)


        # ワークブック（.xlsx）ファイルへのパス
        wb_file_path = GameTreeWorkbookFilePaths.as_workbook(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

        # TODO ワークブック読込
        wb = xl.load_workbook(filename=wb_file_path)

        # TODO ワークブックの Tree シート読込

        # TODO ワークブックに Leafs シート追加
        wb.create_sheet(title='Leafs')

        # TODO Tree シートの葉要素を Leafs シートへ一覧
        tree_ws = wb['Tree']
        for row_no in range(1, ws.max_row + 1):
            for column_no in range(1, ws.max_column+ 1):
                # 最後の有効な列のテキストを葉ノードテキストとみなす
                column_letter = xl.utils.get_column_letter(column_number)
                pass

        # TODO Leafs シートを、科目ごとに集計

        # TODO ワークブックを保存


        print(f"[{datetime.datetime.now()}] Please look {wb_file_path}")


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
