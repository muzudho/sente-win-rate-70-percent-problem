#
# 分析
# python simulate_game_tree_wb.py
#
#   GB を GBWB へ変換します
#

import traceback
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side

from library import HEAD, TAIL, ALICE, IN_GAME, ALICE_FULLY_WON, BOB_FULLY_WON, ALICE_POINTS_WON, BOB_POINTS_WON, NO_WIN_MATCH, Specification, SeriesRule
from library.file_paths import GameTreeFilePaths
from library.database import GameTreeNode, GameTreeRecord, GameTreeTable
from library.workbooks import GameTreeWorkbookWrapper
from library.views import stringify_csv_of_score_board_view_body, PromptCatalog
from library.score_board import search_all_score_boards
from library.views import ScoreBoardViewData
from scripts import SaveOrIgnore


# pts 欄は印を入れるのにも使ってる
PTS_MARK_SAME_RATE = -2


class Prefetch():


    def __init__(self, gt_table_1, gt_table_2):
        self._gt_table_1 = gt_table_1
        self._gt_table_2 = gt_table_2 # prefetched
        self._prev_gt1_record = None


    @staticmethod
    def instantiate(spec, span, t_step, h_step):
        # GTテーブル
        gt_table_1, gt1_file_read_result = GameTreeTable.from_csv(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD),
                new_if_it_no_exists=False)

        if gt1_file_read_result.is_file_not_found:
            raise ValueError(f"GTファイルが見つかりません {gt1_file_read_result.file_path=}")


        # 空のGTテーブルを用意
        gt_table_2 = GameTreeTable.new_empty_table(
                spec=spec,
                span=span,
                t_step=t_step,
                h_step=h_step)

        return Prefetch(gt_table_1=gt_table_1, gt_table_2=gt_table_2)


    @property
    def gt_table_1(self):
        return self._gt_table_1


    @property
    def gt_table_2(self):
        return self._gt_table_2


    def on_gt1_record(self, row_number, gt1_record):
        """
        Parameters
        ----------
        gt1_record : GameTreeRecord
            変更対象のレコード
        """

        # 先頭行は無条件追加
        # -----------------
        if self._prev_gt1_record is None:
            # そのまんま追加
            self._gt_table_2.upsert_record(
                    welcome_record=gt1_record)


        else:

            # リプレース後のレコード。何も更新しなければコピーを返します
            gt2_record = gt1_record.update()


            def same_node_as_avobe(gt1_record, node_no):
                """指定のノードは、上行の繰り返しか？ ただしレートが入っていないノードは常に偽とする"""

                if node_no == 1:
                    prev_nd = self._prev_gt1_record.node1
                    nd = gt1_record.node1
                elif node_no == 2:
                    prev_nd = self._prev_gt1_record.node2
                    nd = gt1_record.node2
                elif node_no == 3:
                    prev_nd = self._prev_gt1_record.node3
                    nd = gt1_record.node3
                elif node_no == 4:
                    prev_nd = self._prev_gt1_record.node4
                    nd = gt1_record.node4
                elif node_no == 5:
                    prev_nd = self._prev_gt1_record.node5
                    nd = gt1_record.node5
                elif node_no == 6:
                    prev_nd = self._prev_gt1_record.node6
                    nd = gt1_record.node6
                else:
                    raise ValueError(f"未対応のノード番号 {node_no=}")

                # レートが入っていなければ偽
                if pd.isnull(nd.rate):
                    return False
                
                # コインの出目と、確率が上行と同じ
                return nd.face == prev_nd.face and nd.rate == prev_nd.rate


            # 1局後
            # -----
            i = 1
            nd = gt1_record.node1

            # TODO セルに上行と同じ値が入っていたら、"├"、"└"、空欄のいずれかにする。ひとまず pts に PTS_MARK_SAME_RATE=-2 を入れておく
            if same_node_as_avobe(gt1_record=gt1_record, node_no=i):
                print(f"[{datetime.datetime.now()}] {gt1_record.no}行目 {i}局後 SAME")
                gt2_record = gt2_record.update(
                        node1=GameTreeNode(
                                face=nd.face,
                                winner=nd.winner,
                                pts=PTS_MARK_SAME_RATE,
                                rate=nd.rate))
                self._gt_table_2.upsert_record(
                        welcome_record=gt2_record)


            # 2局後
            # -----
            i = 2
            nd = gt1_record.node2

            if same_node_as_avobe(gt1_record=gt1_record, node_no=i):
                print(f"[{datetime.datetime.now()}] {gt1_record.no}行目 {i}局後 SAME")
                gt2_record = gt2_record.update(
                        node2=GameTreeNode(
                                face=nd.face,
                                winner=nd.winner,
                                pts=PTS_MARK_SAME_RATE,
                                rate=nd.rate))
                update = self._gt_table_2.upsert_record(
                        welcome_record=gt2_record)
                if not update:
                    raise ValueError(f"アップデートしないのはおかしい {i=}")


            # 3局後
            # -----
            i = 3
            nd = gt1_record.node3

            if same_node_as_avobe(gt1_record=gt1_record, node_no=i):
                print(f"[{datetime.datetime.now()}] {gt1_record.no}行目 {i}局後 SAME")
                gt2_record = gt2_record.update(
                        node3=GameTreeNode(
                                face=nd.face,
                                winner=nd.winner,
                                pts=PTS_MARK_SAME_RATE,
                                rate=nd.rate))
                update = self._gt_table_2.upsert_record(
                        welcome_record=gt2_record)
                if not update:
                    raise ValueError(f"アップデートしないのはおかしい {i=}")


            # 4局後
            # -----
            i = 4
            nd = gt1_record.node4

            if same_node_as_avobe(gt1_record=gt1_record, node_no=i):
                print(f"[{datetime.datetime.now()}] {gt1_record.no}行目 {i}局後 SAME")
                gt2_record = gt2_record.update(
                        node4=GameTreeNode(
                                face=nd.face,
                                winner=nd.winner,
                                pts=PTS_MARK_SAME_RATE,
                                rate=nd.rate))
                update = self._gt_table_2.upsert_record(
                        welcome_record=gt2_record)
                if not update:
                    raise ValueError(f"アップデートしないのはおかしい {i=}")


            # 5局後
            # -----
            i = 5
            nd = gt1_record.node5

            if same_node_as_avobe(gt1_record=gt1_record, node_no=i):
                print(f"[{datetime.datetime.now()}] {gt1_record.no}行目 {i}局後 SAME")
                gt2_record = gt2_record.update(
                        node5=GameTreeNode(
                                face=nd.face,
                                winner=nd.winner,
                                pts=PTS_MARK_SAME_RATE,
                                rate=nd.rate))
                update = self._gt_table_2.upsert_record(
                        welcome_record=gt2_record)
                if not update:
                    raise ValueError(f"アップデートしないのはおかしい {i=}")


            # 6局後
            # -----
            i = 6
            nd = gt1_record.node6

            if same_node_as_avobe(gt1_record=gt1_record, node_no=i):
                print(f"[{datetime.datetime.now()}] {gt1_record.no}行目 {i}局後 SAME")
                gt2_record = gt2_record.update(
                        node6=GameTreeNode(
                                face=nd.face,
                                winner=nd.winner,
                                pts=PTS_MARK_SAME_RATE,
                                rate=nd.rate))
                update = self._gt_table_2.upsert_record(
                        welcome_record=ft2_record)
                if not update:
                    raise ValueError(f"アップデートしないのはおかしい {i=}")


        self._prev_gt1_record = gt1_record


class Automation():


    def __init__(self, gt_table_2, gt_wb_wrapper):
        self._gt_table_2 = gt_table_2
        self._gt_wb_wrapper = gt_wb_wrapper
        self._prev_gt2_record = None


    def on_header(self):

        # 変数名の短縮
        ws = self._gt_wb_wrapper.worksheet


        # 列の幅設定
        # width はだいたい 'ＭＳ Ｐゴシック' サイズ11 の半角英文字の個数

        # TODO C列には確率を入れたい
        # TODO D列は空列にしたい
        # TODO E列の上の方の行には 1 を入れたい

        ws.column_dimensions['A'].width = 4     # no
        ws.column_dimensions['B'].width = 20    # result
        ws.column_dimensions['C'].width = 14    # rate
        ws.column_dimensions['D'].width = 14    # empty column
        ws.column_dimensions['E'].width = 14    # root node
        ws.column_dimensions['F'].width = 2    # 1
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 2    # 2
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 2    # 3
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 10
        ws.column_dimensions['O'].width = 2    # 4
        ws.column_dimensions['P'].width = 14
        ws.column_dimensions['Q'].width = 10
        ws.column_dimensions['R'].width = 2    # 5
        ws.column_dimensions['S'].width = 14
        ws.column_dimensions['T'].width = 10
        ws.column_dimensions['U'].width = 2    # 6
        ws.column_dimensions['V'].width = 14
        ws.column_dimensions['W'].width = 10


        # 行の高さ設定
        # height の単位はポイント。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
        ws.row_dimensions[1].height = 13
        ws.row_dimensions[2].height = 13


        # １行目
        # ------
        # ヘッダー行にする
        row_number = 1

        # インデックス
        column_number = 1
        ws[f'{xl.utils.get_column_letter(column_number)}{row_number}'] = 'no'

        # そのままコピーできない
        # # ２列目～
        # for column_number, column_name in enumerate(self._gt_table_2.df.columns.values, 2):
        #     ws[f'{xl.utils.get_column_letter(column_number)}{row_number}'] = column_name
        ws[f'{xl.utils.get_column_letter(1)}{row_number}'] = 'No'
        ws[f'{xl.utils.get_column_letter(2)}{row_number}'] = '結果'
        ws[f'{xl.utils.get_column_letter(3)}{row_number}'] = '実現確率'
        # 4 は空列

        ws[f'{xl.utils.get_column_letter(5)}{row_number}'] = '開始前'

        # 6 は分岐線
        # 7 はedge
        ws[f'{xl.utils.get_column_letter(8)}{row_number}'] = '1局後'   # node
        ws[f'{xl.utils.get_column_letter(11)}{row_number}'] = '2局後'
        ws[f'{xl.utils.get_column_letter(14)}{row_number}'] = '3局後'
        ws[f'{xl.utils.get_column_letter(17)}{row_number}'] = '4局後'
        ws[f'{xl.utils.get_column_letter(20)}{row_number}'] = '5局後'
        ws[f'{xl.utils.get_column_letter(23)}{row_number}'] = '6局後'

        # ２行目
        # ------
        # 空行にする
        row_number = 2


    def on_each_gt2_record(self, row_number, gt2_record):

        # 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
        node_bgcolor = PatternFill(patternType='solid', fgColor='FFFFCC')

        # 罫線
        side = Side(style='thick', color='000000')
        # style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
        upside_node_border = Border(top=side, left=side, right=side)
        downside_node_border = Border(bottom=side, left=side, right=side)
        under_border = Border(bottom=side)


        # 変数名短縮
        ws = self._gt_wb_wrapper.worksheet


        # ３行目～６行目
        # -------------
        # データは３行目から、１かたまり３行を使って描画する
        rn1 = row_number * 3 + 3
        rn2 = row_number * 3 + 3 + 1
        rn3 = row_number * 3 + 3 + 2
        three_row_numbers = [rn1, rn2, rn3]

        # 行の高さ設定
        # height の単位はポイント。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
        ws.row_dimensions[rn1].height = 13
        ws.row_dimensions[rn2].height = 13
        ws.row_dimensions[rn3].height = 6

        ws[f'A{rn1}'].value = gt2_record.no
        ws[f'B{rn1}'].value = gt2_record.result


        # TODO C列には確率を入れたい。あとで入れる
        # TODO D列は空列にしたい
        # TODO E列の上の方の行には 1 を入れたい


        def draw_node(nd, three_column_names, three_row_numbers):

            if pd.isnull(nd.face):
                print(f"[{datetime.datetime.now()}] face が空欄のノードは無視")
                return

            def edge_text(node):
                if node.face == 'h':
                    face = '表'
                elif node.face == 't':
                    face = '裏'
                elif node.face == 'f':
                    face = '失敗'
                else:
                    raise ValueError(f"{node.face=}")
                
                if node.winner == 'A':
                    winner = '(Ａさん'
                elif node.winner == 'B':
                    winner = '(Ｂさん'
                elif node.winner == 'N':
                    winner = ''
                else:
                    raise ValueError(f"{node.winner=}")

                if node.pts != -1:
                    pts = f"{node.pts:.0f}点)" # FIXME 小数部を消してる。これで誤差で丸めを間違えるケースはあるか？
                else:
                    pts = ''

                return f"{face}{winner}{pts}"

            cn1 = three_column_names[0]
            cn2 = three_column_names[1]
            cn3 = three_column_names[2]
            rn1 = three_row_numbers[0]
            rn2 = three_row_numbers[1]
            rn3 = three_row_numbers[2]

            if nd.face == 'h':
                ws[f'{cn1}{rn1}'].border = under_border

            ws[f'{cn2}{rn1}'].value = edge_text(node=nd)
            ws[f'{cn2}{rn1}'].border = under_border
            ws[f'{cn3}{rn1}'].value = nd.rate
            ws[f'{cn3}{rn1}'].fill = node_bgcolor
            ws[f'{cn3}{rn1}'].border = upside_node_border
            ws[f'{cn3}{rn2}'].fill = node_bgcolor
            ws[f'{cn3}{rn2}'].border = downside_node_border


        # 開始ノード
        # --------
        if rn1 == 3:
            ws[f'E{rn1}'].value = 1
            ws[f'E{rn1}'].fill = node_bgcolor
            ws[f'E{rn1}'].border = upside_node_border
            ws[f'E{rn2}'].fill = node_bgcolor
            ws[f'E{rn2}'].border = downside_node_border

            draw_node(nd=gt2_record.node1, three_column_names=['F', 'G', 'H'], three_row_numbers=three_row_numbers)
            draw_node(nd=gt2_record.node2, three_column_names=['I', 'J', 'K'], three_row_numbers=three_row_numbers)
            draw_node(nd=gt2_record.node3, three_column_names=['L', 'M', 'N'], three_row_numbers=three_row_numbers)
            draw_node(nd=gt2_record.node4, three_column_names=['O', 'P', 'Q'], three_row_numbers=three_row_numbers)
            draw_node(nd=gt2_record.node5, three_column_names=['R', 'S', 'T'], three_row_numbers=three_row_numbers)
            draw_node(nd=gt2_record.node6, three_column_names=['U', 'V', 'W'], three_row_numbers=three_row_numbers)


        else:
            # 実現確率
            rate = None


            # 1局後
            # -----
            i = 1
            nd = gt2_record.node1
            # NOTE 空欄にすべきところには、プリフェッチ時に pts に -2 を入れてある
            if not pd.isnull(nd.pts) and nd.pts != PTS_MARK_SAME_RATE:
                print(f"[{datetime.datetime.now()}] {gt2_record.no}行目 {i}局後 not same")
                draw_node(nd=nd, three_column_names=['F', 'G', 'H'], three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 2局後
            # -----
            i = 2
            nd = gt2_record.node2
            if not pd.isnull(nd.pts) and nd.pts != PTS_MARK_SAME_RATE:
                print(f"[{datetime.datetime.now()}] {gt2_record.no}行目 {i}局後 not same")
                draw_node(nd=nd, three_column_names=['I', 'J', 'K'], three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 3局後
            # -----
            i = 3
            nd = gt2_record.node3
            if not pd.isnull(nd.pts) and nd.pts != PTS_MARK_SAME_RATE:
                print(f"[{datetime.datetime.now()}] {gt2_record.no}行目 {i}局後 not same")
                draw_node(nd=nd, three_column_names=['L', 'M', 'N'], three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 4局後
            # -----
            i = 4
            nd = gt2_record.node4
            if not pd.isnull(nd.pts) and nd.pts != PTS_MARK_SAME_RATE:
                print(f"[{datetime.datetime.now()}] {gt2_record.no}行目 {i}局後 not same")
                draw_node(nd=nd, three_column_names=['O', 'P', 'Q'], three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 5局後
            # -----
            i = 5
            nd = gt2_record.node5
            if not pd.isnull(nd.pts) and nd.pts != PTS_MARK_SAME_RATE:
                print(f"[{datetime.datetime.now()}] {gt2_record.no}行目 {i}局後 not same")
                draw_node(nd=nd, three_column_names=['R', 'S', 'T'], three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 6局後
            # -----
            i = 6
            nd = gt2_record.node6
            if not pd.isnull(nd.pts) and nd.pts != PTS_MARK_SAME_RATE:
                print(f"[{datetime.datetime.now()}] {gt2_record.no}行目 {i}局後 not same")
                draw_node(nd=nd, three_column_names=['U', 'V', 'W'], three_row_numbers=three_row_numbers)

            if not pd.isnull(nd.rate):
                rate = nd.rate


            # 実現確率
            # --------
            ws[f'C{rn1}'].value = rate


        self._prev_gt2_record = gt2_record


########################################
# コマンドから実行時
########################################
if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # ［先後の決め方］を尋ねます
        specified_turn_system_id = PromptCatalog.which_method_do_you_use_to_determine_sente_and_gote()


        # ［将棋の引分け率］を尋ねます
        specified_failure_rate = PromptCatalog.what_is_the_failure_rate()


        # ［将棋の先手勝率］を尋ねます
        specified_p = PromptCatalog.what_is_the_probability_of_flipping_a_coin_and_getting_heads()


        # ［目標の点数］を尋ねます
        specified_span = PromptCatalog.how_many_goal_win_points()


        # ［後手で勝ったときの勝ち点］を尋ねます
        specified_t_step = PromptCatalog.how_many_win_points_of_tail_of_coin()


        # ［先手で勝ったときの勝ち点］を尋ねます
        specified_h_step = PromptCatalog.how_many_win_points_of_head_of_coin()


        # ［仕様］
        spec = Specification(
                turn_system_id=specified_turn_system_id,
                failure_rate=specified_failure_rate,
                p=specified_p)

        # FIXME 便宜的に［試行シリーズ数］は 1 固定
        specified_trial_series = 1

        # ［シリーズ・ルール］。任意に指定します
        specified_series_rule = SeriesRule.make_series_rule_base(
                spec=spec,
                span=specified_span,
                t_step=specified_t_step,
                h_step=specified_h_step)


        prefetch = Prefetch.instantiate(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

        # GTテーブルをプリフェッチする
        # TODO １回シート全体を舐めて樹形のアテンションを加える必要があるか？ "├" とか "└" のアテンション
        prefetch.gt_table_1.for_each(on_each=prefetch.on_gt1_record)

        print(f"""\
[{datetime.datetime.now()}] prefetch.gt_table_2.df:
{prefetch.gt_table_2.df}""")


        # GTWB ファイル作成
        # オリジナルはインターフェースに癖があるので、ラッパーを作成してそれを使う
        gt_wb_wrapper = GameTreeWorkbookWrapper.instantiate(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

        # ワークブックを開く（既存なら削除してから新規作成）
        gt_wb = gt_wb_wrapper.open_workbook(remove_workbook_if_it_exists=True)

        # GameTree シートを作成
        gt_wb_wrapper.create_sheet('GameTree', shall_overwrite=False)

        # 既存の Sheet シートを削除
        gt_wb_wrapper.remove_sheet('Sheet')

        automation = Automation(gt_table_2=prefetch.gt_table_2, gt_wb_wrapper=gt_wb_wrapper)

        # GTWB の Sheet シートへのヘッダー書出し
        automation.on_header()

        # GTWB の Sheet シートへの各行書出し
        prefetch.gt_table_2.for_each(on_each=automation.on_each_gt2_record)

        # GTWB ファイルの保存
        gt_wb_wrapper.save()


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
