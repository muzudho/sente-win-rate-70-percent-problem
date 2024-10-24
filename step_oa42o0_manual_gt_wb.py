#
# python step_oa42o0_manual_gt_wb.py
#
#   * `wb` - Workbook
#   GT を GTWB へ変換します
#

import math
import traceback
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
import xltree as tr

from library import HEAD, TAIL, Specification, SeriesRule
from library.file_paths import GameTreeWorkbookFilePaths, GameTreeFilePaths
from library.database import GameTreeNode, GameTreeRecord, GameTreeTable
from library.workbooks import GameTreeWorkbookWrapper
from library.views import PromptCatalog
from library.game_tree_view import GameTreeView


########################################
# コマンドから実行時
########################################
if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # ［先後の決め方］を尋ねます
        specified_turn_system_id = PromptCatalog.which_method_do_you_use_to_determine_sente_and_gote()


        # ［将棋の引分け率］を尋ねます
        specified_failure_rate = PromptCatalog.what_is_the_failure_rate()


        # ［将棋の先手勝率］を尋ねます
        specified_p = PromptCatalog.what_is_the_probability_of_flipping_a_coin_and_getting_heads()


        # ［目標の点数］を尋ねます
        specified_span = PromptCatalog.how_many_goal_win_points()


        # ［後手で勝ったときの勝ち点］を尋ねます
        specified_t_step = PromptCatalog.how_many_win_points_of_tail_of_coin()


        # ［先手で勝ったときの勝ち点］を尋ねます
        specified_h_step = PromptCatalog.how_many_win_points_of_head_of_coin()


        # ［仕様］
        spec = Specification(
                turn_system_id=specified_turn_system_id,
                failure_rate=specified_failure_rate,
                p=specified_p)

        # FIXME 便宜的に［試行シリーズ数］は 1 固定
        specified_trial_series = 1

        # ［シリーズ・ルール］。任意に指定します
        specified_series_rule = SeriesRule.make_series_rule_base(
                spec=spec,
                span=specified_span,
                t_step=specified_t_step,
                h_step=specified_h_step)


        # ワークブック（.xlsx）ファイルへのパス
        wb_file_path = GameTreeWorkbookFilePaths.as_workbook(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

        # 構成
        settings ={
            # # 列の幅
            # 'no_width':                         4,      # A列の幅。no列
            # 'row_header_separator_width':       3,      # B列の幅。空列
            # 'node_width':                       20,     # 例：C, F, I ...列の幅。ノードの箱の幅
            # 'parent_side_edge_width':           2,      # 例：D, G, J ...列の幅。エッジの水平線のうち、親ノードの方
            # 'child_side_edge_width':            4,      # 例：E, H, K ...列の幅。エッジの水平線のうち、子ノードの方

            # # 行の高さ
            # 'header_height':                    13,     # 第１行。ヘッダー
            # 'column_header_separator_height':   13,     # 第２行。空行
        }

        # 出力先ワークブックを指定し、ワークブックハンドル取得
        with tr.prepare_workbook(target=wb_file_path, mode='w', settings=settings) as b:

            csv_file_path=GameTreeFilePaths.as_csv(
                    spec=spec,
                    span=specified_series_rule.step_table.span,
                    t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                    h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

            # テーブル読取
            df = pd.read_csv(csv_file_path, encoding="utf8", index_col=['no'])
            #print(df)

            # 読取元CSVを指定し、ワークシートハンドル取得
            with b.prepare_worksheet(target='Tree', based_on=csv_file_path) as s:

                # ワークシートへ木構造図を描画
                s.render_tree()

                # TODO 集計を行いたい。ツリー構造の全ての葉を取得する                
                all_leaf_nodes = []

                def search(all_leaf_nodes, node):
                    """再帰的に子ノードを表示"""
                    for child_node in node.child_nodes.values():
                        # 葉ノード
                        if len(child_node.child_nodes) < 1:
                            all_leaf_nodes.append(child_node)
                        
                        # 中間ノード
                        else:
                            search(all_leaf_nodes=all_leaf_nodes, node=child_node) # 再帰

                for root_node in s.multiple_root_node.values():
                    search(all_leaf_nodes=all_leaf_nodes, node=root_node)
                                
                # leaf_th と result列 を紐づける
                rate_list_by_result = {}
                for leaf in all_leaf_nodes:
                    result = df.at[leaf.leaf_th, 'result']
                    #print(f"葉テスト {leaf.text=}  {leaf.leaf_th=}  {result}")

                    if result not in rate_list_by_result:
                        rate_list_by_result[result] = []
                    
                    rate_list_by_result[result].append(float(leaf.text))

                # result 別に確率を高精度 sum する
                sum_rate_by_result = {}
                for result, rate_list in rate_list_by_result.items():
                    sum_rate = math.fsum(rate_list)
                    #print(f"{result=}  {sum_rate=}")
                    sum_rate_by_result[result] = sum_rate

                # 結果表示 ＆ Total検算
                total = math.fsum(sum_rate_by_result.values())
                #print(f"検算 {total=}")

                if total != 1:
                    raise ValueError(f"total must be 1. but {total}")

            # 読取元CSVを指定し、ワークシートハンドル取得
            with b.prepare_worksheet(target='Summary', based_on=csv_file_path) as s:
                ws = s._ws  # 非公式な方法。将来的にサポートされるか分からない方法

                # データフレームの操作
                # ------------------

                # 操作が便利なので、pandas に移し替える
                records = []
                for result, sum_rate in sum_rate_by_result.items():
                    records.append([result, sum_rate])

                df = pd.DataFrame(records, columns=['result', 'sum_rate'])

                # sum_rate 列の値の大きい順に並び変える
                df.sort_values(['sum_rate', 'result'], inplace=True, ascending=[False, True])

                # 連番列を追加
                df['no'] = range(0, len(df))

                # 連番列を、（列を止めて）インデックスに変更
                df.set_index('no', inplace=True)

                # 列の並び替え
                #df = df[['result', 'sum_rate']]

                # デバッグ表示
                print(df)


                # 最長の文字数も図っておく
                if len(df) == 0:
                    max_length_of_a = 0
                    max_length_of_b = 0
                else:
                    max_length_of_a = df['result'].str.len().max()  # 文字列の長さ
                    max_length_of_b = df['sum_rate'].abs().astype(str).str.len().max()-1    # 浮動小数点数の長さ


                # Total を集計
                total_sum_rate = df['sum_rate'].sum()


                # ワークシートへの出力
                # ------------------

                # 列名
                ws['A1'] = 'result'
                ws['B1'] = 'sum_rate'

                # 列幅の自動調整
                ws.column_dimensions['A'].width = max_length_of_a * 2       # 日本語
                ws.column_dimensions['B'].width = max_length_of_b * 1.2     # 浮動小数点数

                # ヘッダーの背景色
                fill = PatternFill(patternType='solid', fgColor='111111')
                ws['A1'].fill = fill
                ws['B1'].fill = fill

                # ヘッダーの文字色
                font = Font(color='EEEEEE')
                ws['A1'].font = font
                ws['B1'].font = font

                # データ部のコピー
                for row_no in range(0, len(df)):
                    result = df.at[row_no, 'result']
                    sum_rate = df.at[row_no, 'sum_rate']
                    # データ部は 2nd 行目から
                    ws[f'A{row_no + 2}'] = result
                    ws[f'B{row_no + 2}'] = sum_rate
                    ws[f'B{row_no + 2}'].alignment = Alignment(horizontal='left')

                # データ部のトータル行の追加
                row_no += 1
                ws[f'A{row_no + 2}'] = 'Total'
                ws[f'B{row_no + 2}'] = total_sum_rate
                ws[f'B{row_no + 2}'].alignment = Alignment(horizontal='left')

                # データ部のトータル行を区切る罫線
                side = Side(style='thick', color='111111')
                border = Border(top=side)
                ws[f'A{row_no + 2}'].border = border
                ws[f'B{row_no + 2}'].border = border


            # 何かワークシートを１つ作成したあとで、最初から入っている 'Sheet' を削除
            b.remove_worksheet(target='Sheet')

            # 保存
            b.save_workbook()

            print(f"[{datetime.datetime.now()}] Please look {wb_file_path}")


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
