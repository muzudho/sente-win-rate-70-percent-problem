#
# python simulate_gt_wb.py
#
#   * `wb` - Workbook
#   GT を GTWB へ変換します
#

import math
import traceback
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
import xltree as tr

from library import HEAD, TAIL, Specification, SeriesRule
from library.file_paths import GameTreeWorkbookFilePaths, GameTreeFilePaths
from library.database import GameTreeNode, GameTreeRecord, GameTreeTable
from library.workbooks import GameTreeWorkbookWrapper
from library.views import PromptCatalog
from library.game_tree_view import GameTreeView


########################################
# コマンドから実行時
########################################
if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # ［先後の決め方］を尋ねます
        specified_turn_system_id = PromptCatalog.which_method_do_you_use_to_determine_sente_and_gote()


        # ［将棋の引分け率］を尋ねます
        specified_failure_rate = PromptCatalog.what_is_the_failure_rate()


        # ［将棋の先手勝率］を尋ねます
        specified_p = PromptCatalog.what_is_the_probability_of_flipping_a_coin_and_getting_heads()


        # ［目標の点数］を尋ねます
        specified_span = PromptCatalog.how_many_goal_win_points()


        # ［後手で勝ったときの勝ち点］を尋ねます
        specified_t_step = PromptCatalog.how_many_win_points_of_tail_of_coin()


        # ［先手で勝ったときの勝ち点］を尋ねます
        specified_h_step = PromptCatalog.how_many_win_points_of_head_of_coin()


        # ［仕様］
        spec = Specification(
                turn_system_id=specified_turn_system_id,
                failure_rate=specified_failure_rate,
                p=specified_p)

        # FIXME 便宜的に［試行シリーズ数］は 1 固定
        specified_trial_series = 1

        # ［シリーズ・ルール］。任意に指定します
        specified_series_rule = SeriesRule.make_series_rule_base(
                spec=spec,
                span=specified_span,
                t_step=specified_t_step,
                h_step=specified_h_step)


        # ワークブック（.xlsx）ファイルへのパス
        wb_file_path = GameTreeWorkbookFilePaths.as_workbook(
                spec=spec,
                span=specified_series_rule.step_table.span,
                t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

        # 構成
        settings ={
            # # 列の幅
            # 'no_width':                         4,      # A列の幅。no列
            # 'row_header_separator_width':       3,      # B列の幅。空列
            # 'node_width':                       20,     # 例：C, F, I ...列の幅。ノードの箱の幅
            # 'parent_side_edge_width':           2,      # 例：D, G, J ...列の幅。エッジの水平線のうち、親ノードの方
            # 'child_side_edge_width':            4,      # 例：E, H, K ...列の幅。エッジの水平線のうち、子ノードの方

            # # 行の高さ
            # 'header_height':                    13,     # 第１行。ヘッダー
            # 'column_header_separator_height':   13,     # 第２行。空行
        }

        # 出力先ワークブックを指定し、ワークブックハンドル取得
        with tr.prepare_workbook(target=wb_file_path, mode='w', settings=settings) as b:

            csv_file_path=GameTreeFilePaths.as_csv(
                    spec=spec,
                    span=specified_series_rule.step_table.span,
                    t_step=specified_series_rule.step_table.get_step_by(face_of_coin=TAIL),
                    h_step=specified_series_rule.step_table.get_step_by(face_of_coin=HEAD))

            # テーブル読取
            df = pd.read_csv(csv_file_path, encoding="utf8", index_col=['no'])
            #print(df)

            # 読取元CSVを指定し、ワークシートハンドル取得
            with b.prepare_worksheet(target='Tree', based_on=csv_file_path) as s:

                # ワークシートへ木構造図を描画
                s.render_tree()

                # TODO 集計を行いたい。ツリー構造の全ての葉を取得する                
                all_leaf_nodes = []

                def search(all_leaf_nodes, node):
                    """再帰的に子ノードを表示"""
                    for child_node in node.child_nodes.values():
                        # 葉ノード
                        if len(child_node.child_nodes) < 1:
                            all_leaf_nodes.append(child_node)
                        
                        # 中間ノード
                        else:
                            search(all_leaf_nodes=all_leaf_nodes, node=child_node) # 再帰

                for root_node in s.multiple_root_node.values():
                    search(all_leaf_nodes=all_leaf_nodes, node=root_node)
                                
                # leaf_th と result列 を紐づける
                rate_list_by_result = {}
                for leaf in all_leaf_nodes:
                    result = df.at[leaf.leaf_th, 'result']
                    #print(f"葉テスト {leaf.text=}  {leaf.leaf_th=}  {result}")

                    if result not in rate_list_by_result:
                        rate_list_by_result[result] = []
                    
                    rate_list_by_result[result].append(float(leaf.text))

                # result 別に確率を高精度 sum する
                sum_rate_by_result = {}
                for result, rate_list in rate_list_by_result.items():
                    sum_rate = math.fsum(rate_list)
                    #print(f"{result=}  {sum_rate=}")
                    sum_rate_by_result[result] = sum_rate

                # 結果表示 ＆ Total検算
                total = math.fsum(sum_rate_by_result.values())
                #print(f"検算 {total=}")

                if total != 1:
                    raise ValueError(f"total must be 1. but {total}")

            # 読取元CSVを指定し、ワークシートハンドル取得
            with b.prepare_worksheet(target='Summary', based_on=csv_file_path) as s:
                ws = s._ws  # 非公式な方法。将来的にサポートされるか分からない方法

                # 列名
                ws['A1'] = 'result'
                ws['B1'] = sum_rate

                for row_th, (result, sum_rate) in enumerate(sum_rate_by_result.items(), 2):
                    ws[f'A{row_th}'] = result
                    ws[f'B{row_th}'] = sum_rate


            # 何かワークシートを１つ作成したあとで、最初から入っている 'Sheet' を削除
            b.remove_worksheet(target='Sheet')

            # 保存
            b.save_workbook()

            print(f"[{datetime.datetime.now()}] Please look {wb_file_path}")


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
