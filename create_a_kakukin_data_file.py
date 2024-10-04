#
# TODO 作成中
#
# python create_a_kakukin_data_file.py
#
# Excel ファイルを作ろう
#
import traceback
import openpyxl as xl
import os

from library import FROZEN_TURN, ALTERNATING_TURN, Converter, Specification
from library.file_paths import get_kakukin_data_excel_file_path, get_kakukin_data_sheet_csv_file_path
from library.database import KakukinDataSheetTable


ws = None
row_number = 0


########################################
# コマンドから実行時
########################################
if __name__ == '__main__':
    try:
        # ［将棋の引分け率］を尋ねる
        prompt = f"""\
What is the failure rate?
Example: 10% is 0.1
? """
        specified_failure_rate = float(input(prompt))


        # ［先後の決め方］を尋ねる
        prompt = f"""\
(1) Frozen turn
(2) Alternating turn
Which one(1-2)? """

        choice = input(prompt)
        if choice == '1':
            specified_turn_system = FROZEN_TURN
        elif choice == '2':
            specified_turn_system = ALTERNATING_TURN
        else:
            raise ValueError(f"{choice=}")


        # ［試行シリーズ数］を尋ねる
        prompt = f"""\
How many times do you want to try the series?

(0) Try       2 series
(1) Try      20 series
(2) Try     200 series
(3) Try    2000 series
(4) Try   20000 series
(5) Try  200000 series
(6) Try 2000000 series

Example: 3
(0-6)? """
        precision = int(input(prompt))
        specified_trials_series = Converter.precision_to_trials_series(precision)
        specified_abs_small_error = Converter.precision_to_small_error(precision)


        excel_file_path = get_kakukin_data_excel_file_path(turn_system=specified_turn_system, trials_series=specified_trials_series)


        # ワークブックの作成
        wb = xl.Workbook()

        while True:
            if os.path.isfile(excel_file_path):
                command = input(f"""\
{excel_file_path} という名前のファイルは既にあります。
上書きしますか(Y/n)? """)

                if command == 'n':
                    continue

            break

        wb.save(excel_file_path)

        print(f"""\
{excel_file_path} を保存しました。
""")

        sheet_name = "f0"

        # 最初に Sheet という名前のシートができているので、それを参照します
        ws = wb["Sheet"]

        # シートの名前を変更します
        ws.title = sheet_name
        wb.save(excel_file_path)

        # 例えば `KDS_alter_f0.0_try2000.csv` といったファイルの内容を、シートに移していきます
        # 📖 [openpyxlで別ブックにシートをコピーする](https://qiita.com/github-nakasho/items/fb9df8e423bb8784cbbd)

        df_kds = KakukinDataSheetTable.read_df(
                failure_rate=specified_failure_rate,
                turn_system=specified_turn_system,
                trials_series=specified_trials_series)


        print(f"""\
シートの名前を {ws.title} に変更しました。
{excel_file_path} を保存しました。
""")

        # 列名 A ～
        column_names = [
            'p',
            'failure_rate',
            'turn_system',
            'head_step',
            'tail_step',
            'span',
            'shortest_coins',
            'upper_limit_coins',
            'trials_series',
            'series_shortest_coins',
            'series_longest_coins',
            'wins_a',
            'wins_b',
            'succucessful_series',
            's_ful_wins_a',
            's_ful_wins_b',
            's_pts_wins_a',
            's_pts_wins_b',
            'failed_series',
            'f_ful_wins_a',
            'f_ful_wins_b',
            'f_pts_wins_a',
            'f_pts_wins_b',
            'no_wins_ab',
        ]

        for index, column_name in enumerate(column_names, 1):
            ws[f'{xl.utils.get_column_letter(index)}1'] = column_name

        # データ部
        # --------

        row_number = 2

        def on_each(record):
            global ws, row_number

            ws[f'A{row_number}'].value = record.p
            ws[f'B{row_number}'].value = record.failure_rate
            ws[f'C{row_number}'].value = record.turn_system
            ws[f'D{row_number}'].value = record.head_step
            ws[f'E{row_number}'].value = record.tail_step
            ws[f'F{row_number}'].value = record.span
            ws[f'G{row_number}'].value = record.shortest_coins
            ws[f'H{row_number}'].value = record.upper_limit_coins
            ws[f'I{row_number}'].value = record.trials_series
            ws[f'J{row_number}'].value = record.series_shortest_coins
            ws[f'K{row_number}'].value = record.series_longest_coins
            ws[f'L{row_number}'].value = record.wins_a
            ws[f'M{row_number}'].value = record.wins_b
            ws[f'N{row_number}'].value = record.succucessful_series
            ws[f'O{row_number}'].value = record.s_ful_wins_a
            ws[f'P{row_number}'].value = record.s_ful_wins_b
            ws[f'Q{row_number}'].value = record.s_pts_wins_a
            ws[f'R{row_number}'].value = record.s_pts_wins_b
            ws[f'S{row_number}'].value = record.failed_series
            ws[f'T{row_number}'].value = record.f_ful_wins_a
            ws[f'U{row_number}'].value = record.f_ful_wins_b
            ws[f'V{row_number}'].value = record.f_pts_wins_a
            ws[f'W{row_number}'].value = record.f_pts_wins_b
            ws[f'X{row_number}'].value = record.no_wins_ab

            row_number += 1

        KakukinDataSheetTable.for_each(
                df=df_kds,
                on_each=on_each)


        wb.save(excel_file_path)

        print(f"""でーきたっ！""")


    except Exception as err:
        print(f"""\
おお、残念！　例外が投げられてしまった！  
{err=}  {type(err)=}

以下はスタックトレース表示じゃ。
{traceback.format_exc()}
""")
