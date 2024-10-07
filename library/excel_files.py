import os
import openpyxl as xl


from library.file_paths import KakukinDataFilePaths


class KakukinDataExcelFile():
    """［かくきんデータ・エクセル・ファイル］"""


    def __init__(self, excel_file_path):
        """初期化"""
        self._excel_file_path = excel_file_path
        self._wb = None
        self._current_ws = None


    @staticmethod
    def instantiate(turn_system_id, trial_series):
        """インスタンス生成"""

        # エクセル・ファイルへのパス
        excel_file_path = KakukinDataFilePaths.as_excel(
                turn_system_id=turn_system_id,
                trial_series=trial_series)

        return KakukinDataExcelFile(
                excel_file_path=excel_file_path)


    def load_workbook(self):
        # ファイルが既存なら読込
        if os.path.isfile(self._excel_file_path):
            self._wb = xl.load_workbook(filename=self._excel_file_path)

        # ファイルが存在しなければ新規作成
        else:
            # ワークブックの作成
            self._wb = xl.Workbook()
        
        return self._wb


    def create_sheet(self, title, shall_overwrite=False):
        """シート作成

        Parameters
        ----------
        title : str
            シートの名前
        """

        # もしシートが既存なら削除する
        if shall_overwrite:
            if title in self._wb.sheetnames:
                del self._wb[title]

        self._current_ws = self._wb.create_sheet(title=title)
        return self._current_ws


    def save(self):
        """エクセル・ファイルの保存
        
        Returns
        -------
        excel_file_path : str
            ファイルパス
        """
        self._wb.save(self._excel_file_path)
        return self._excel_file_path
