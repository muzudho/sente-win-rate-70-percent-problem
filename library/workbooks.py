import os
import openpyxl as xl


from library.file_paths import GameTreeWorkbookFilePaths


class GameTreeWorkbookWrapper():
    """［ゲームツリー］ワークブック（.xlsx）"""


    def __init__(self, wb_file_path):
        """初期化"""
        self._wb_file_path = wb_file_path
        self._wb = None
        self._ws = None


    @staticmethod
    def instantiate(spec, span, t_step, h_step):
        """インスタンス生成"""

        # ワークブック（.xlsx）ファイルへのパス
        wb_file_path = GameTreeWorkbookFilePaths.as_workbook(
                spec=spec,
                span=span,
                t_step=t_step,
                h_step=h_step)

        return GameTreeWorkbookWrapper(
                wb_file_path=wb_file_path)


    @property
    def workbook(self):
        """このワークブック"""
        return self._wb


    @property
    def worksheet(self):
        """現在作業中のシート"""
        return self._ws


    def open_workbook(self, remove_workbook_if_it_exists):
        """ワークブック（.xlsx）ファイルを開く"""
        
        # ファイルが既存なら
        if os.path.isfile(self._wb_file_path):
            # 既存のファイルを削除して新規作成
            if remove_workbook_if_it_exists:
                # 削除前の安全策
                if not self._wb_file_path.endswith('.xlsx'):
                    raise ValueError(f"エクセル形式のファイルが指定されていません。 {self._wb_file_path=}")

                try:               
                    os.remove(self._wb_file_path)
                
                # FIXME エクセルファイルが開けっ放しのとき
                # PermissionError: [WinError 32] プロセスはファイルにアクセスできません。別のプロセスが使用中です。: 'reports/kakukin/auto_generated_kakukin_data_try2000_alter.xlsx'
                except PermissionError as e:
                    raise PermissionError('既存のエクセルファイルを削除できませんでした。作業をスキップします') from e

                shall_new = True

            # 読込
            else:
                shall_new = False

                try:
                    self._wb = xl.load_workbook(filename=self._wb_file_path)

                except KeyError as e:
                    print(f"""\
{e}
{self._wb_file_path=}
""")
                    raise
        
        # ファイルが存在しなければ新規作成
        else:
            shall_new = True


        # 新規作成
        if shall_new:
            # ワークブックの作成
            self._wb = xl.Workbook()


        return self._wb


    def create_sheet(self, sheet_name, shall_overwrite):
        """シート作成

        Parameters
        ----------
        sheet_name : str
            シートの名前
        """

        # もしシートが既存なら削除する
        if shall_overwrite:
            if sheet_name in self._wb.sheetnames:
                del self._wb[sheet_name]

        self._ws = self._wb.create_sheet(title=sheet_name) # 既存だとコピーを作ってしまう？
        return self._ws


    def remove_sheet(self, sheet_name):
        self._wb.remove(self._wb[sheet_name])


    def save(self):
        """ワークブック（.xlsx）ファイルの保存
        
        Returns
        -------
        wb_file_path : str
            ファイルパス
        """
        self._wb.save(self._wb_file_path)
        return self._wb_file_path
