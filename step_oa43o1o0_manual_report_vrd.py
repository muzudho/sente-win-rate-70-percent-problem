#
# python step_oa43o1o0_manual_report_vrd.py
#
#   VRS表を CSV形式から XLSX形式へ変換します
#

import traceback
import datetime
import os
import random
import math
import time
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

from library import SeriesRule, get_list_of_basename
from library.file_basename import BasenameOfVictoryRateDetailFile
from library.file_paths import VictoryRateDetailFilePaths


########################################
# コマンドから実行時
########################################


if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # victory_rate_detail フォルダーを見る
        detail_basename_list = get_list_of_basename(dir_path=VictoryRateDetailFilePaths.get_temp_directory_path())

        # シャッフル
        random.shuffle(detail_basename_list)


        for detail_basename in detail_basename_list:

            # ファイル名から［仕様］を取得する
            spec = BasenameOfVictoryRateDetailFile.to_spec(basename=detail_basename)

            if spec is None:
                continue

            try:

                # ファイルパス取得
                # ---------------
                detail_csv_file_path = f'./{VictoryRateDetailFilePaths.get_temp_directory_path()}/{detail_basename}'    # VRS表 CSV
                detail_wb_file_path = VictoryRateDetailFilePaths.as_workbook_on_reports(spec=spec)                      # ワークブック

                print(f"[{datetime.datetime.now()}] step_oa43o1o0_manual_report_vrd {detail_csv_file_path=}  {detail_wb_file_path=}")
                time.sleep(1)   # １秒休む


                # CSV読取
                # -------
                if os.path.isfile(detail_csv_file_path):   # VRS表(CSV)。既存時
                    df = pd.read_csv(
                            detail_csv_file_path,
                            encoding="utf-8")

                    # t_time が無ければ追加
                    if 't_time' not in df.columns.values:
                        df['t_time'] = (df['span'] / df['t_step']).map(math.ceil)

                    # h_time が無ければ追加
                    if 'h_time' not in df.columns.values:
                        df['h_time'] = (df['span'] / df['h_step']).map(math.ceil)

                    # shortest_coins が無ければ追加
                    if 'shortest_coins' not in df.columns.values:
                        df['shortest_coins'] = df[['span', 't_step', 'h_step']].apply(lambda X:SeriesRule.let_shortest_coins(h_step=X['h_step'], t_step=X['t_step'], span=X['span'], turn_system_id=spec.turn_system_id), axis=1)

                    # upper_limit_coins が無ければ追加
                    if 'upper_limit_coins' not in df.columns.values:
                        df['upper_limit_coins'] = df[['t_time', 'h_time']].apply(lambda X:SeriesRule.let_upper_limit_coins_without_failure_rate(spec=spec, h_time=X['h_time'], t_time=X['t_time']), axis=1)


                    # 列順の変更
                    df = df[[
                        'h_step',                   # 表点
                        't_step',                   # ｳﾗ点
                        'span',                     # 優勝点
                        'h_time',                   # 必要表回数
                        't_time',                   # 必要ｳﾗ回数
                        'shortest_coins',           # 最短対局数
                        'upper_limit_coins',        # 対局数上限
                        'a_victory_rate_by_trio',
                        'b_victory_rate_by_trio',
                        'no_victory_rate',
                        'a_victory_rate_by_duet',
                        'b_victory_rate_by_duet',
                        'unfair_point']]


                    # ソート
                    #
                    #   優先度１： unfair_point が 0 に近く
                    #   優先度２： no_victory_rate が 0 に近く
                    #   優先度３： span が小さく
                    #   優先度４： h_step が小さく      ※ t_step <= h_step なので、長い方を小さくしたい
                    #   優先度５： t_step が小さい
                    #
                    df.sort_values(
                            by=['unfair_point','no_victory_rate', 'span', 'h_step', 't_step'],
                            inplace=True)

                else:
                    raise ValueError(f"file not found {detail_csv_file_path=}")


                # ワークブックの新規作成
                wb = xl.Workbook()

                # 既定シート名変更
                wb['Sheet'].title = 'Detail'

                # シート取得
                ws = wb['Detail']

                # 列幅設定
                #
                #   数値通りにはいかない
                #   浮動小数点数にキリがないので、列名の長さに揃えるのも手
                #
                column_width_list = [
                    ('A', 4.64),    # 表点
                    ('B', 5.00),    # ｳﾗ点
                    ('C', 6.64),    # 優勝点
                    ('D', 10.73),   # 必要表回数
                    ('E', 11.18),   # 必要ｳﾗ回数
                    ('F', 10.73),   # 最短対局数
                    ('G', 10.73),   # 対局数上限
                    ('H', 5.0),     # ［Ａさんの優勝率（優勝なし率込）］
                    ('I', 5.0),     # ［Ｂさんの優勝率（優勝なし率込）］
                    ('J', 10.0),    # ［優勝なし率］
                    ('K', 14.00),   # ［Ａさんの優勝率］
                    ('L', 14.00),   # ［Ｂさんの優勝率］
                    ('M', 7.5),     # ［不均等度］
                ]
                for pair in column_width_list:
                    ws.column_dimensions[pair[0]].width = pair[1]

                # 列名の変更
                new_column_name_dict = {
                    'h_step':'表点',
                    't_step':'ｳﾗ点',
                    'span':'優勝点',
                    'h_time':'必要表回数',
                    't_time':'必要ｳﾗ回数',
                    'shortest_coins':'最短対局数',
                    'upper_limit_coins':'対局数上限',
                    'a_victory_rate_by_trio':'Ａさんの優勝率（優勝なし率込）',
                    'b_victory_rate_by_trio':'Ｂさんの優勝率（優勝なし率込）',
                    'no_victory_rate':'優勝なし率',
                    'a_victory_rate_by_duet':'Ａさんの優勝率',
                    'b_victory_rate_by_duet':'Ｂさんの優勝率',
                    'unfair_point':'不均等度'}

                # 寄せ
                alignment_list = [
                    Alignment(horizontal='right'),  # 表点
                    Alignment(horizontal='right'),  # ｳﾗ点
                    Alignment(horizontal='right'),  # 優勝点
                    Alignment(horizontal='right'),  # 必要表回数
                    Alignment(horizontal='right'),  # 必要ｳﾗ回数
                    Alignment(horizontal='right'),  # 最短対局数
                    Alignment(horizontal='right'),  # 対局数上限
                    Alignment(horizontal='left'),   # Ａさんの優勝率（優勝なし率込）
                    Alignment(horizontal='left'),   # Ｂさんの優勝率（優勝なし率込）
                    Alignment(horizontal='left'),   # 優勝なし率
                    Alignment(horizontal='left'),   # Ａさんの優勝率
                    Alignment(horizontal='left'),   # Ｂさんの優勝率
                    Alignment(horizontal='left'),   # 不均等度
                ]

                # ヘッダー文字色・背景色
                header_font = Font(color='EEFFEE')
                header_background_fill = PatternFill(patternType='solid', fgColor='336633')

                # ヘッダー出力
                for column_th, column_name in enumerate(df.columns.values, 1):
                    column_letter = xl.utils.get_column_letter(column_th)
                    cell = ws[f'{column_letter}1']
                    cell.value = new_column_name_dict[column_name]
                    cell.font = header_font
                    cell.fill = header_background_fill
                    cell.alignment = alignment_list[column_th - 1]


                # データ部
                # -------
                column_letter_and_column_name = [
                    ('A', 'h_step'),                    # 表点
                    ('B', 't_step'),                    # ｳﾗ点
                    ('C', 'span'),                      # 優勝点
                    ('D', 'h_time'),                    # 必要表回数
                    ('E', 't_time'),                    # 必要ｳﾗ回数
                    ('F', 'shortest_coins'),            # 最短対局数
                    ('G', 'upper_limit_coins'),         # 対局数上限
                    ('H', 'a_victory_rate_by_trio'),
                    ('I', 'b_victory_rate_by_trio'),
                    ('J', 'no_victory_rate'),
                    ('K', 'a_victory_rate_by_duet'),
                    ('L', 'b_victory_rate_by_duet'),
                    ('M', 'unfair_point'),
                ]

                for sorted_row_no, (row_no, row) in enumerate(df.iterrows()):
                    #print(f"{row_th=} {type(row)=}")
                    row_th = sorted_row_no + 2

                    for column_no in range(0, len(row)):
                        cell = ws[f'{column_letter_and_column_name[column_no][0]}{row_th}']
                        cell.value = row[column_letter_and_column_name[column_no][1]]
                        cell.alignment = alignment_list[column_no]


                # ウィンドウ枠の固定
                ws.freeze_panes = 'A2'


                # データ部背景色
                # -------------
                light_red_fill = PatternFill(patternType='solid', fgColor='FFEEEE')
                light_yellow_fill = PatternFill(patternType='solid', fgColor='FFFFEE')


                # データ部罫線出力
                # ---------------
                black_side = Side(style='thick', color='333333')
                left_border = Border(left=black_side)
                right_border = Border(right=black_side)
                top_border = Border(top=black_side)

                for row_th in range(2, ws.max_row + 1):

                    # ［表点］
                    cell = ws[f'A{row_th}']
                    cell.border = left_border
                    cell.fill = light_red_fill

                    # ［ｳﾗ点］
                    cell = ws[f'B{row_th}']
                    cell.fill = light_red_fill

                    # ［優勝点］
                    cell = ws[f'C{row_th}']
                    cell.border = right_border
                    cell.fill = light_red_fill

                    # ［Ａさんの優勝率（優勝なし率込）］
                    cell = ws[f'H{row_th}']
                    cell.border = left_border

                    # ［Ａさんの優勝率］
                    cell = ws[f'K{row_th}']
                    cell.border = left_border
                    cell.fill = light_yellow_fill

                    # ［Ｂさんの優勝率］
                    cell = ws[f'L{row_th}']
                    cell.border = right_border
                    cell.fill = light_yellow_fill


                # ワークブックの保存
                wb.save(detail_wb_file_path)
                print(f"[{datetime.datetime.now()}] please look `{detail_wb_file_path}`")


            except PermissionError as e:
                message = f"[{datetime.datetime.now()}] ファイルが他で開かれているのかも？ {detail_csv_file_path=} {detail_wb_file_path=} {e=}"
                print(message)
                # スタックトレース表示
                print(traceback.format_exc())

                log_file_path = VictoryRateDetailFilePaths.as_log(spec=spec)
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"{message}\n")    # ファイルへ出力

                # １分休む
                seconds = 60
                print(f"[{datetime.datetime.now()}] retry after {seconds} seconds")
                time.sleep(seconds)


            except Exception as e:
                message = f"[{datetime.datetime.now()}] 予期せぬ例外 {detail_csv_file_path=} {detail_wb_file_path=} {e=}"
                print(message)
                # スタックトレース表示
                print(traceback.format_exc())

                log_file_path = VictoryRateDetailFilePaths.as_log(spec=spec)
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"{message}\n")    # ファイルへ出力

                # １分休む
                seconds = 60
                print(f"[{datetime.datetime.now()}] retry after {seconds} seconds")
                time.sleep(seconds)


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
