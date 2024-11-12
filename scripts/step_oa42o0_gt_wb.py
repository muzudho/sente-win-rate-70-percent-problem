import math
import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
import xltree as tr

from library import HEAD, TAIL
from library.file_paths import GameTreeWorkbookFilePaths, GameTreeFilePaths, GameTreeCheckedFilePaths


class GeneratorOfGTWB():


    @staticmethod
    def instantiate(series_rule):
        # ファイルパス取得
        # ---------------
        wb_file_path = GameTreeWorkbookFilePaths.as_workbook(   # 出力するワークブック（.xlsx）ファイルへのパス
                spec=series_rule.spec,
                span=series_rule.span,
                t_step=series_rule.get_step_by(face_of_coin=TAIL),
                h_step=series_rule.get_step_by(face_of_coin=HEAD))

        source_csv_file_path=GameTreeFilePaths.as_csv(  # 元となる CSVファイルパス
                spec=series_rule.spec,
                span=series_rule.span,
                t_step=series_rule.get_step_by(face_of_coin=TAIL),
                h_step=series_rule.get_step_by(face_of_coin=HEAD))

        checked_csv_file_path=GameTreeCheckedFilePaths.as_csv(  # チェック済みの CSVファイルパス
                spec=series_rule.spec,
                span=series_rule.span,
                t_step=series_rule.get_step_by(face_of_coin=TAIL),
                h_step=series_rule.get_step_by(face_of_coin=HEAD))

        return GeneratorOfGTWB(series_rule=series_rule, wb_file_path=wb_file_path, source_csv_file_path=source_csv_file_path, checked_csv_file_path=checked_csv_file_path)


    def __init__(self, series_rule, wb_file_path, source_csv_file_path, checked_csv_file_path):
        self._series_rule = series_rule
        self._wb_file_path = wb_file_path
        self._source_csv_file_path = source_csv_file_path
        self._checked_csv_file_path = checked_csv_file_path


    @property
    def workbook_file_path(self):
        """出力するワークブック（.xlsx）ファイルへのパス"""
        return self._wb_file_path


    @property
    def source_csv_file_path(self):
        """元となる CSVファイルパス"""
        return self._source_csv_file_path


    @property
    def checked_csv_file_path(self):
        """チェック済みの CSVファイルパス"""
        return self._checked_csv_file_path


    def write_workbook(self, debug_write=False):
        """ワークブック（.xlsx）ファイルを書き出す

        TODO 高速化したい
        """


        # xltree 用構成
        settings ={
            # # 列の幅
            # 'no_width':                         4,      # A列の幅。no列
            # 'row_header_separator_width':       3,      # B列の幅。空列
            # 'node_width':                       20,     # 例：C, F, I ...列の幅。ノードの箱の幅
            # 'parent_side_edge_width':           2,      # 例：D, G, J ...列の幅。エッジの水平線のうち、親ノードの方
            # 'child_side_edge_width':            4,      # 例：E, H, K ...列の幅。エッジの水平線のうち、子ノードの方

            # # 行の高さ
            # 'header_height':                    13,     # 第１行。ヘッダー
            # 'column_header_separator_height':   13,     # 第２行。空行
        }

        # 出力先ワークブックを指定し、ワークブックハンドル取得
        with tr.prepare_workbook(target=self._wb_file_path, mode='w', settings=settings) as b:

            # テーブル読取
            df = pd.read_csv(self._source_csv_file_path, encoding="utf8", index_col=['no'])
            if debug_write:
                print(df)


            ############
            # MARK: Tree
            ############

            # 読取元CSVを指定し、ワークシートハンドル取得
            with b.prepare_worksheet(target='Tree', based_on=self._source_csv_file_path) as s:

                # ワークシートへ木構造図を描画
                s.render_tree()

                # 集計を行いたい。ツリー構造の全ての葉を取得する                
                all_leaf_entries = []

                def search(all_leaf_entries, entry):
                    """再帰的に子ノードを表示"""

                    if debug_write:
                        print(f"{entry.edge_text=}  {entry.node_text=}  の子要素数={len(entry.child_entries)}  TOTAL {len(all_leaf_entries)=}")

                    for child_entry in entry.child_entries.values():
                        # 葉ノード
                        if not child_entry.has_children():
                            if debug_write:
                                print("葉ノード")
                            all_leaf_entries.append(child_entry)
                        
                        # 中間ノード
                        else:
                            if debug_write:
                                print("中間ノード")
                            search(all_leaf_entries=all_leaf_entries, entry=child_entry) # 再帰

                if debug_write:
                    # 木構造の簡易ターミナル表示
                    print(s.forest._stringify_like_tree(''))
                        

                # 各根について
                for root_entry in s.forest.multiple_root_entry.values():
                    search(all_leaf_entries=all_leaf_entries, entry=root_entry)


                if debug_write:
                    print(f"葉の探索結果  {len(all_leaf_entries)=}")


                # leaf_th と result列 を紐づける
                rate_list_by_result = {}
                for leaf in all_leaf_entries:
                    result = df.at[leaf.leaf_th, 'result']
                    if debug_write:
                        print(f"葉テスト {leaf.node_text=}  {leaf.leaf_th=}  {result}")

                    if result not in rate_list_by_result:
                        rate_list_by_result[result] = []
                    
                    rate_list_by_result[result].append(float(leaf.node_text))

                # result 別に確率を高精度 sum する
                sum_rate_by_result = {}
                for result, rate_list in rate_list_by_result.items():
                    sum_rate = math.fsum(rate_list)
                    if debug_write:
                        print(f"{result=}  {sum_rate=}")
                    sum_rate_by_result[result] = sum_rate

                # 結果表示 ＆ Total検算
                total = math.fsum(sum_rate_by_result.values())
                if debug_write:
                    print(f"検算 {total=}")

                # エラーは常時表示するが、続行する
                if total != 1:
                    print(f"[error] total must be 1. but {total}")
                    #raise ValueError(f"total must be 1. but {total}")


            ###############
            # MARK: Summary
            ###############

            # 読取元CSVを指定し、ワークシートハンドル取得
            with b.prepare_worksheet(target='Summary', based_on=self._source_csv_file_path) as s:
                ws = s._ws  # 非公式な方法。将来的にサポートされるか分からない方法

                # データフレームの操作
                # ------------------

                # 操作が便利なので、pandas に移し替える
                records = []
                for result, sum_rate in sum_rate_by_result.items():
                    records.append([result, sum_rate])

                df = pd.DataFrame(records, columns=['result', 'sum_rate'])

                # sum_rate 列の値の大きい順に並び変える
                df.sort_values(['sum_rate', 'result'], inplace=True, ascending=[False, True])

                # 連番列を追加
                df['no'] = range(0, len(df))

                # 連番列を、（列を止めて）インデックスに変更
                df.set_index('no', inplace=True)

                # 列の並び替え
                #df = df[['result', 'sum_rate']]

                # デバッグ表示
                #print(df)


                # ワークシートへの出力
                # ------------------

                # 列幅
                # ----
                ws.column_dimensions['A'].width = 50.45     # ［Ｂさん（シリーズをｳﾗで始めた方）が優勝した率（失敗込）］ が 50.45
                ws.column_dimensions['B'].width = 12.0      # だいたい

                # 1st ヘッダー行
                self.render_header(df=df, ws=ws, destination_row_th=1)

                # 2nd
                a_victory_rate_by_duet = self.get_a_victory_rate_by_duet(sum_rate_by_result=sum_rate_by_result)
                self.render_item(ws=ws, destination_row_th=2, name='Ａさん（シリーズを表で始めた方）が優勝した率（失敗抜）', value=a_victory_rate_by_duet, alignment=Alignment(horizontal='left'))

                # 3rd
                b_victory_rate_by_duet = self.get_b_victory_rate_by_duet(sum_rate_by_result=sum_rate_by_result)
                self.render_item(ws=ws, destination_row_th=3, name='Ｂさん（シリーズをｳﾗで始めた方）が優勝した率（失敗抜）', value=b_victory_rate_by_duet, alignment=Alignment(horizontal='left'))

                # 4th Total
                self.render_item_with_top_border(ws=ws, destination_row_th=4, name='Total', value=math.fsum([a_victory_rate_by_duet, b_victory_rate_by_duet]))

                # 5th 空行

                # 6th ヘッダー行
                self.render_header(df=df, ws=ws, destination_row_th=6)

                # 7th
                a_victory_rate_by_trio = self.get_a_victory_rate_by_trio(sum_rate_by_result=sum_rate_by_result)
                self.render_item(ws=ws, destination_row_th=7, name='Ａさん（シリーズを表で始めた方）が優勝した率（失敗込）', value=a_victory_rate_by_trio, alignment=Alignment(horizontal='left'))

                # 8th
                b_victory_rate_by_trio = self.get_b_victory_rate_by_trio(sum_rate_by_result=sum_rate_by_result)
                self.render_item(ws=ws, destination_row_th=8, name='Ｂさん（シリーズをｳﾗで始めた方）が優勝した率（失敗込）', value=b_victory_rate_by_trio, alignment=Alignment(horizontal='left'))

                # 9th
                no_victory_rate = self.get_no_victory_rate(sum_rate_by_result=sum_rate_by_result)
                self.render_item(ws=ws, destination_row_th=9, name='優勝が決まらなかった率（失敗）', value=no_victory_rate, alignment=Alignment(horizontal='left'))

                # 10th トータル行
                self.render_item_with_top_border(ws=ws, destination_row_th=10, name='Total', value=math.fsum([a_victory_rate_by_trio, b_victory_rate_by_trio, no_victory_rate]))

                # 11th 空行

                # 12th ヘッダー行
                self.render_header(df=df, ws=ws, destination_row_th=12)

                # 13th～: データ部のコピー
                for source_row_no in range(0, len(df)):
                    destination_row_th = source_row_no + 13

                    self.render_sections(
                            df=df,
                            ws=ws,
                            source_row_no=source_row_no,
                            destination_row_th=destination_row_th)


                # 詳細トータル行
                self.render_item_with_top_border(ws=ws, destination_row_th=destination_row_th + 1, name='Total', value=df['sum_rate'].sum())


            # 何かワークシートを１つ作成したあとで、最初から入っている 'Sheet' を削除
            b.remove_worksheet(target='Sheet')

            # 保存
            b.save_workbook()

            print(f"[{datetime.datetime.now()}] Please look {self._wb_file_path}")


    def render_header(self, df, ws, destination_row_th):
        """ヘッダーを出力"""

        # 列名
        ws[f'A{destination_row_th}'] = 'name'
        ws[f'B{destination_row_th}'] = 'value'

        # ヘッダーの背景色
        fill = PatternFill(patternType='solid', fgColor='111111')
        ws[f'A{destination_row_th}'].fill = fill
        ws[f'B{destination_row_th}'].fill = fill

        # ヘッダーの文字色
        font = Font(color='EEEEEE')
        ws[f'A{destination_row_th}'].font = font
        ws[f'B{destination_row_th}'].font = font


    def render_item(self, ws, destination_row_th, name, value, alignment):
        """表の項目を出力"""

        ws[f'A{destination_row_th}'].value = name
        ws[f'B{destination_row_th}'].value = value
        ws[f'B{destination_row_th}'].alignment = alignment


    def get_a_victory_rate_by_trio(self, sum_rate_by_result):
        """［Ａさん（シリーズを表で始めた方）が優勝した率（失敗込）］を算出"""
        # result 別に確率を高精度 sum する
        rate_list = []

        if '達成でＡさんの勝ち' in sum_rate_by_result:
            rate_list.append(sum_rate_by_result['達成でＡさんの勝ち'])
        
        if '勝ち点差でＡさんの勝ち' in sum_rate_by_result:
            rate_list.append(sum_rate_by_result['勝ち点差でＡさんの勝ち'])

        return math.fsum(rate_list)


    def get_a_victory_rate_by_duet(self, sum_rate_by_result):
        """［Ａさん（シリーズを表で始めた方）が優勝した率（失敗抜）］を算出"""
        a_victory_rate_by_trio = self.get_a_victory_rate_by_trio(sum_rate_by_result=sum_rate_by_result)

        no_victory_rate = 0
        if '勝者なし' in sum_rate_by_result:
            no_victory_rate = sum_rate_by_result['勝者なし']

        if no_victory_rate == 1:
            return 0

        return a_victory_rate_by_trio / (1 - no_victory_rate)


    def get_b_victory_rate_by_trio(self, sum_rate_by_result):
        """［Ｂさん（シリーズをｳﾗで始めた方）が優勝した率（失敗込）］を算出"""

        # result 別に確率を高精度 sum する
        rate_list = []

        if '達成でＢさんの勝ち' in sum_rate_by_result:
            rate_list.append(sum_rate_by_result['達成でＢさんの勝ち'])
        
        if '勝ち点差でＢさんの勝ち' in sum_rate_by_result:
            rate_list.append(sum_rate_by_result['勝ち点差でＢさんの勝ち'])

        return math.fsum(rate_list)


    def get_b_victory_rate_by_duet(self, sum_rate_by_result):
        """［Ｂさん（シリーズをｳﾗで始めた方）が優勝した率（失敗抜）］を算出"""
        b_victory_rate_by_trio = self.get_b_victory_rate_by_trio(sum_rate_by_result=sum_rate_by_result)

        no_victory_rate = 0
        if '勝者なし' in sum_rate_by_result:
            no_victory_rate = sum_rate_by_result['勝者なし']

        if no_victory_rate == 1:
            return 0

        return b_victory_rate_by_trio / (1 - no_victory_rate)


    def get_no_victory_rate(self, sum_rate_by_result):
        """［優勝が決まらなかった率（失敗）］を算出"""

        if '勝者なし' in sum_rate_by_result:
            return sum_rate_by_result['勝者なし']
        
        return 0


    def render_item_with_top_border(self, ws, destination_row_th, name, value):
        """上線付き項目を出力"""

        # データ部のトータル行の追加
        ws[f'A{destination_row_th}'] = name
        ws[f'B{destination_row_th}'] = value
        ws[f'B{destination_row_th}'].alignment = Alignment(horizontal='left')

        # データ部のトータル行を区切る罫線
        side = Side(style='thick', color='111111')
        border = Border(top=side)
        ws[f'A{destination_row_th}'].border = border
        ws[f'B{destination_row_th}'].border = border


    def render_sections(self, df, ws, source_row_no, destination_row_th):
        result = df.at[source_row_no, 'result']
        sum_rate = df.at[source_row_no, 'sum_rate']

        ws[f'A{destination_row_th}'] = result
        ws[f'B{destination_row_th}'] = sum_rate
        ws[f'B{destination_row_th}'].alignment = Alignment(horizontal='left')


    def render_detail_total(self, df, ws, destination_row_th, value):
        """詳細トータル行を出力"""

        # データ部のトータル行の追加
        ws[f'A{destination_row_th}'] = 'Total'
        ws[f'B{destination_row_th}'] = value
        ws[f'B{destination_row_th}'].alignment = Alignment(horizontal='left')

        # データ部のトータル行を区切る罫線
        side = Side(style='thick', color='111111')
        border = Border(top=side)
        ws[f'A{destination_row_th}'].border = border
        ws[f'B{destination_row_th}'].border = border
