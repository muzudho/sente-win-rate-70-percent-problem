#
# ［かくきんデータ］Excel ファイルを作ろう
#
import traceback
import openpyxl as xl
import os
import datetime

from library import FROZEN_TURN, ALTERNATING_TURN, Converter, Specification
from library.file_paths import KakukinDataFilePaths
from library.database import KakukinDataSheetTable
from library.excel_files import KakukinDataExcelFile


class Automation():
    """自動化"""


    def __init__(self, specified_failure_rate, specified_turn_system_id, specified_trial_series):
        """初期化

        Parameters
        ----------
        specified_failure_rate : float
            ［コインを投げて表も裏も出ない確率］
        specified_turn_system_id : int
            ［先後の決め方］
        specified_trial_series : int
            ［試行シリーズ数］
        """
        self._specified_failure_rate = specified_failure_rate
        self._specified_turn_system_id = specified_turn_system_id
        self._specified_trial_series = specified_trial_series

        self._ws = None
        self._row_number = 0


    def on_each(self, record):
        self._ws[f'A{self._row_number}'].value = record.p
        self._ws[f'B{self._row_number}'].value = record.failure_rate
        self._ws[f'C{self._row_number}'].value = record.turn_system_name

        self._ws[f'D{self._row_number}'].value = record.head_step               # TODO ［シリーズ・ルール］は、理論値が選ばれるように仕様変更したい
        self._ws[f'E{self._row_number}'].value = record.tail_step
        self._ws[f'F{self._row_number}'].value = record.span
        self._ws[f'G{self._row_number}'].value = record.shortest_coins
        self._ws[f'H{self._row_number}'].value = record.upper_limit_coins

        self._ws[f'I{self._row_number}'].value = record.trial_series           # TODO ［シミュレーション結果］は、理論値の［シリーズ・ルール］をもとに、計測し直したい
        self._ws[f'J{self._row_number}'].value = record.series_shortest_coins   
        self._ws[f'K{self._row_number}'].value = record.series_longest_coins    
        self._ws[f'L{self._row_number}'].value = record.wins_a
        self._ws[f'M{self._row_number}'].value = record.wins_b
        self._ws[f'N{self._row_number}'].value = record.succucessful_series
        self._ws[f'O{self._row_number}'].value = record.s_ful_wins_a
        self._ws[f'P{self._row_number}'].value = record.s_ful_wins_b
        self._ws[f'Q{self._row_number}'].value = record.s_pts_wins_a
        self._ws[f'R{self._row_number}'].value = record.s_pts_wins_b
        self._ws[f'S{self._row_number}'].value = record.failed_series
        self._ws[f'T{self._row_number}'].value = record.f_ful_wins_a
        self._ws[f'U{self._row_number}'].value = record.f_ful_wins_b
        self._ws[f'V{self._row_number}'].value = record.f_pts_wins_a
        self._ws[f'W{self._row_number}'].value = record.f_pts_wins_b
        self._ws[f'X{self._row_number}'].value = record.no_wins_ab

        self._row_number += 1


    def execute(self):
        """実行
        
        NOTE 先にKDSファイルを作成しておく必要があります
        """

        # 対エクセル・ファイル用オブジェクト作成
        kakukin_data_excel_file = KakukinDataExcelFile.instantiate(
                turn_system_id=self._specified_turn_system_id,
                trial_series=self._specified_trial_series)

        # エクセル・ファイルの読込
        kakukin_data_excel_file.load_workbook()

        # シートの名前を作成するぞ（シートが既存なら上書き）
        #
        #   Example: ［将棋の引分け率］が 0.05 なら `f5.0per`
        #   NOTE シート名に "%" を付けると Excel の式が動かなくなった
        #
        sheet_name = f'f{self._specified_failure_rate * 100:.1f}per'
        self._ws = kakukin_data_excel_file.create_sheet(title=sheet_name, shall_overwrite=True)


        # 例えば `KDS_alter_f0.0_try2000.csv` といったファイルの内容を、シートに移していきます
        # 📖 [openpyxlで別ブックにシートをコピーする](https://qiita.com/github-nakasho/items/fb9df8e423bb8784cbbd)

        kds_table = KakukinDataSheetTable.read_csv(
                failure_rate=self._specified_failure_rate,
                turn_system_id=self._specified_turn_system_id,
                trial_series=self._specified_trial_series)


        # KDSファイルが無かったのならスキップする
        if kds_table.df is None:
            print(f"[{datetime.datetime.now()}] KDSファイルが無かったのならスキップする")
            return


        # ヘッダー部
        # ----------
        for index, column_name in enumerate(kds_table.df.columns.values, 1):
            self._ws[f'{xl.utils.get_column_letter(index)}1'] = column_name

        # データ部
        # --------
        self._row_number = 2

        kds_table.for_each(on_each=self.on_each)


        # ［かくきんデータ・エクセル・ファイル］保存
        excel_file_path = kakukin_data_excel_file.save()
        print(f"[{datetime.datetime.now()}] saved: `{excel_file_path}` file")
