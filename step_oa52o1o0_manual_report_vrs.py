#
# python step_oa52o1o0_manual_report_vrs.py
#
#   VRS表を CSV形式から XLSX形式へ変換します
#

import traceback
import datetime
import os
import random
import math
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

from library import HEAD, TAIL, Specification, Converter, SeriesRule, toss_a_coin
from library.file_paths import VictoryRateSummaryFilePaths


########################################
# コマンドから実行時
########################################


if __name__ == '__main__':
    """コマンドから実行時"""

    try:

        # ファイルパス取得
        # ---------------
        summary_csv_file_path = VictoryRateSummaryFilePaths.as_csv()                # VRS表 CSV
        summary_wb_file_path = VictoryRateSummaryFilePaths.as_workbook_on_reports() # ワークブック

        # CSV読取
        # -------
        if os.path.isfile(summary_csv_file_path):   # VRS表(CSV)。既存時
            summary_df = pd.read_csv(
                    summary_csv_file_path,
                    encoding="utf-8")
        
            # 列順の変更
            summary_df = summary_df[[
                'p',                        # 先手勝率
                'failure_rate',             # 引分け率
                'turn_system_name',         # 手番の決め方
                'h_step',                   # 表点
                't_step',                   # ｳﾗ点
                'span',                     # 優勝点
                'a_victory_rate_by_trio',
                'b_victory_rate_by_trio',
                'no_victory_rate',
                'unfair_point']]

            # ソート
            summary_df.sort_values(
                    by=['p','failure_rate','turn_system_name'],
                    ascending=[
                        True,
                        True,
                        False], # 固定→交互の順
                    inplace=True)

        else:
            raise ValueError(f"file not found {summary_csv_file_path=}")


        # ワークブックの新規作成
        wb = xl.Workbook()

        # 既定シート名変更
        wb['Sheet'].title = 'Summary'

        # シート取得
        ws = wb['Summary']

        # 出力する列の設計
        #
        #   大した量ではないので、辞書をネストさせる
        #
        output_column_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
        output_columns = {
            # KEY:列番号
            # Value:
            #   'label' - 表示列名
            #   'name' - 元データ列名
            #   'width' - 列幅設定
            #                   数値通りにはいかない
            #                   浮動小数点数にキリがないので、列名の長さに揃えるのも手
            #   'align' - 位置寄せ
            #
            'A': {'label':'先手勝率', 'name':'p', 'width':8.64, 'align':Alignment(horizontal='left')},
            'B': {'label':'引分率', 'name':'failure_rate', 'width':6.64, 'align':Alignment(horizontal='left')},
            'C': {'label':'手番の決め方', 'name':'turn_system_name', 'width':12.73, 'align':Alignment(horizontal='left')},
            'D': {'label':'表点', 'name':'h_step', 'width':4.64, 'align':Alignment(horizontal='right')},
            'E': {'label':'ｳﾗ点', 'name':'t_step', 'width':5.00, 'align':Alignment(horizontal='right')},
            'F': {'label':'優勝点', 'name':'span', 'width':6.64, 'align':Alignment(horizontal='right')},
            'G': {'label':'必要表回数', 'name':None, 'width':10.73, 'align':Alignment(horizontal='right')},     # 追加 'h_time'
            'H': {'label':'必要ｳﾗ回数', 'name':None, 'width':11.18, 'align':Alignment(horizontal='right')},     # 追加 't_time'
            'I': {'label':'最短対局数', 'name':None, 'width':10.73, 'align':Alignment(horizontal='right')},     # 追加 'shortest_coins'
            'J': {'label':'対局数上限', 'name':None, 'width':10.73, 'align':Alignment(horizontal='right')},     # 追加 'upper_limit_coins'
            'K': {'label':'Ａさんの優勝率（優勝なし率込）', 'name':'a_victory_rate_by_trio', 'width':5.0, 'align':Alignment(horizontal='left')},
            'L': {'label':'Ｂさんの優勝率（優勝なし率込）', 'name':'b_victory_rate_by_trio', 'width':5.0, 'align':Alignment(horizontal='left')},
            'M': {'label':'優勝なし率', 'name':'no_victory_rate', 'width':16.55, 'align':Alignment(horizontal='left')},
            'N': {'label':'Ａさんの優勝率', 'name':None, 'width':13.91, 'align':Alignment(horizontal='left')},      # 追加 'a_victory_rate_by_duet'
            'O': {'label':'Ｂさんの優勝率', 'name':None, 'width':14.00, 'align':Alignment(horizontal='left')},      # 追加 'b_victory_rate_by_duet'
            'P': {'label':'不均等度', 'name':'unfair_point', 'width':8.5, 'align':Alignment(horizontal='left')}
        }

        for column_letter in output_column_letters:
            ws.column_dimensions[column_letter].width = output_columns[column_letter]['width']
        

        # ヘッダー文字色・背景色
        header_font = Font(color='EEFFEE')
        header_background_fill = PatternFill(patternType='solid', fgColor='336633')


        # ヘッダー出力
        for column_letter in output_column_letters:
            cell = ws[f'{column_letter}1']
            cell.value = output_columns[column_letter]['label']
            cell.font = header_font
            cell.fill = header_background_fill
            cell.alignment = output_columns[column_letter]['align']


        turn_system_name_to_jp = {
            'alternating':'交互',
            'frozen':'固定',
        }


        for sorted_row_no, (row_no, row) in enumerate(summary_df.iterrows()):
            #print(f"{row_th=} {type(row)=}")
            row_th = sorted_row_no + 2

            # ［仕様］
            spec = Specification(
                    turn_system_id=Converter.turn_system_code_to_id(row['turn_system_name']),
                    failure_rate=row['failure_rate'],
                    p=row['p'])

            upper_limit_coins_letter = None
            h_time = None
            t_time = None
            for column_letter in output_column_letters:

                cell = ws[f'{column_letter}{row_th}']
                name = output_columns[column_letter]['name']

                # 元データにある列
                if name is not None:
                    # ［手番の決め方］の値を日本語化
                    if name == 'turn_system_name':
                        cell.value = turn_system_name_to_jp[row[name]]
                    
                    else:
                        cell.value = row[name]

                # 元データにない列                
                else:
                    label = output_columns[column_letter]['label']

                    if label == '必要表回数':
                        h_time = math.ceil(row['span'] / row['h_step'])
                        cell.value = h_time

                    elif label == '必要ｳﾗ回数':
                        t_time = math.ceil(row['span'] / row['t_step'])
                        cell.value = t_time

                    elif label == '最短対局数':
                        cell.value = SeriesRule.let_shortest_coins(h_step=row['h_step'], t_step=row['t_step'], span=row['span'], turn_system_id=spec.turn_system_id)

                    elif label == '対局数上限':
                        # あとで設定
                        upper_limit_coins_letter = column_letter

                    elif label == 'Ａさんの優勝率':
                        cell.value = row['a_victory_rate_by_trio'] / (1 - row['no_victory_rate'])
                    
                    elif label == 'Ｂさんの優勝率':
                        cell.value = row['b_victory_rate_by_trio'] / (1 - row['no_victory_rate'])

                    else:
                        raise ValueError(f"{label=}")


                cell.alignment = output_columns[column_letter]['align']


            # 対局数上限
            cell = ws[f'{upper_limit_coins_letter}{row_th}']
            cell.value = SeriesRule.let_upper_limit_coins_without_failure_rate(spec=spec, h_time=h_time, t_time=t_time)


        # ウィンドウ枠の固定
        ws.freeze_panes = 'A2'


        # データ部背景色
        # -------------
        light_red_fill = PatternFill(patternType='solid', fgColor='FFEEEE')
        light_green_fill = PatternFill(patternType='solid', fgColor='EEFFEE')
        light_blue_fill = PatternFill(patternType='solid', fgColor='EEEEFF')


        # データ部罫線出力
        # ---------------
        black_side = Side(style='thick', color='333333')
        left_border = Border(left=black_side)
        right_border = Border(right=black_side)
        top_border = Border(top=black_side)

        pre_p = None

        for row_th in range(2, ws.max_row + 1):

            # ［先手勝率］
            cell = ws[f'A{row_th}']
            # ［先手勝率］の変わり目
            if pre_p is not None and pre_p != cell.value:
                cell.border = top_border

            pre_p = cell.value


            # ［表点］
            cell = ws[f'D{row_th}']
            cell.border = left_border
            cell.fill = light_red_fill

            # ［ｳﾗ点］
            cell = ws[f'E{row_th}']
            cell.fill = light_red_fill

            # ［優勝点］
            cell = ws[f'F{row_th}']
            cell.border = right_border
            cell.fill = light_red_fill

            # ［Ａさんの優勝率（優勝なし率込）］
            cell = ws[f'K{row_th}']
            cell.border = left_border

            # ［対局数上限］
            cell = ws[f'J{row_th}']
            cell.fill = light_green_fill

            # ［Ａさんの優勝率］
            cell = ws[f'N{row_th}']
            cell.border = left_border
            cell.fill = light_blue_fill

            # ［Ｂさんの優勝率］
            cell = ws[f'O{row_th}']
            cell.border = right_border
            cell.fill = light_blue_fill


        # ワークブックの保存
        wb.save(summary_wb_file_path)
        print(f"[{datetime.datetime.now()}] please look `{summary_wb_file_path}`")


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
