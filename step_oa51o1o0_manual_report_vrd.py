#
# python step_oa51o1o0_manual_report_vrd.py
#
#   VRS表を CSV形式から XLSX形式へ変換します
#

import traceback
import datetime
import os
import random
import math
import time
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

from library import SeriesRule, get_list_of_basename
from library.file_basename import BasenameOfVictoryRateDetailFile
from library.file_paths import VictoryRateDetailFilePaths


########################################
# コマンドから実行時
########################################


if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # victory_rate_detail フォルダーを見る
        detail_basename_list = get_list_of_basename(dir_path=VictoryRateDetailFilePaths.get_temp_directory_path())

        # シャッフル
        random.shuffle(detail_basename_list)


        for detail_basename in detail_basename_list:

            # ファイル名から［仕様］を取得する
            spec = BasenameOfVictoryRateDetailFile.to_spec(basename=detail_basename)

            if spec is None:
                continue

            try:

                # ファイルパス取得
                # ---------------
                detail_csv_file_path = f'./{VictoryRateDetailFilePaths.get_temp_directory_path()}/{detail_basename}'    # VRS表 CSV
                detail_wb_file_path = VictoryRateDetailFilePaths.as_workbook_on_reports(spec=spec)                      # ワークブック

                print(f"[{datetime.datetime.now()}] step_oa52o1o0_manual_report_vrs {detail_csv_file_path=}  {detail_wb_file_path=}")
                time.sleep(1)   # １秒休む


                # CSV読取
                # -------
                if os.path.isfile(detail_csv_file_path):   # VRS表(CSV)。既存時
                    df = pd.read_csv(
                            detail_csv_file_path,
                            encoding="utf-8")

                    # TODO こういう書き方もある
                    # # upper_limit_coins が無ければ追加
                    # if 'upper_limit_coins' not in df.columns.values:
                    #     df['upper_limit_coins'] = df[['t_time', 'h_time']].apply(lambda X:SeriesRule.let_upper_limit_coins_without_failure_rate(spec=spec, h_time=X['h_time'], t_time=X['t_time']), axis=1)


                    # 列順の変更
                    df = df[[
                        'h_step',                   # 表点
                        't_step',                   # ｳﾗ点
                        'span',                     # 優勝点
                        'a_victory_rate_by_trio',
                        'b_victory_rate_by_trio',
                        'no_victory_rate',
                        'unfair_point']]


                    # ソート
                    #
                    #   優先度１： unfair_point が 0 に近く
                    #   優先度２： no_victory_rate が 0 に近く
                    #   優先度３： span が小さく
                    #   優先度４： h_step が小さく      ※ t_step <= h_step なので、長い方を小さくしたい
                    #   優先度５： t_step が小さい
                    #
                    df.sort_values(
                            by=['unfair_point','no_victory_rate', 'span', 'h_step', 't_step'],
                            inplace=True)

                else:
                    raise ValueError(f"file not found {detail_csv_file_path=}")


                # ワークブックの新規作成
                wb = xl.Workbook()

                # 既定シート名変更
                wb['Sheet'].title = 'Detail'

                # シート取得
                ws = wb['Detail']

                # 出力する列の設計
                #
                #   大した量ではないので、辞書をネストさせる
                #
                output_column_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M']
                output_columns = {
                    # KEY:列番号
                    # Value:
                    #   'label' - 表示列名
                    #   'name' - 元データ列名
                    #   'width' - 列幅設定
                    #                   数値通りにはいかない
                    #                   浮動小数点数にキリがないので、列名の長さに揃えるのも手
                    #   'align' - 位置寄せ
                    #
                    'A': {'label':'表点', 'name':'h_step', 'width':4.64, 'align':Alignment(horizontal='right')},
                    'B': {'label':'ｳﾗ点', 'name':'t_step', 'width':5.00, 'align':Alignment(horizontal='right')},
                    'C': {'label':'優勝点', 'name':'span', 'width':6.64, 'align':Alignment(horizontal='right')},
                    'D': {'label':'必要表回数', 'name':None, 'width':10.73, 'align':Alignment(horizontal='right')},     # 'h_time' 追加
                    'E': {'label':'必要ｳﾗ回数', 'name':None, 'width':11.18, 'align':Alignment(horizontal='right')},     # 't_time' 追加
                    'F': {'label':'最短対局数', 'name':None, 'width':10.73, 'align':Alignment(horizontal='right')},     # 'shortest_coins' 追加
                    'G': {'label':'対局数上限', 'name':None, 'width':10.73, 'align':Alignment(horizontal='right')},     # 'upper_limit_coins' 追加
                    'H': {'label':'Ａさんの優勝率（失敗込）', 'name':'a_victory_rate_by_trio', 'width':5.0, 'align':Alignment(horizontal='left')},
                    'I': {'label':'Ｂさんの優勝率（失敗込）', 'name':'b_victory_rate_by_trio', 'width':5.0, 'align':Alignment(horizontal='left')},
                    'J': {'label':'優勝なし率（失敗）', 'name':'no_victory_rate', 'width':10.0, 'align':Alignment(horizontal='left')},
                    'K': {'label':'Ａさんの優勝率（失敗抜）', 'name':None, 'width':14.0, 'align':Alignment(horizontal='left')},       # 'a_victory_rate_by_duet' 追加
                    'L': {'label':'Ｂさんの優勝率（失敗抜）', 'name':None, 'width':14.0, 'align':Alignment(horizontal='left')},       # 'b_victory_rate_by_duet' 追加
                    'M': {'label':'不均等度', 'name':'unfair_point', 'width':8.5, 'align':Alignment(horizontal='left')},
                }

                for column_letter in output_column_letters:
                    ws.column_dimensions[column_letter].width = output_columns[column_letter]['width']

                # ヘッダー文字色・背景色
                header_font = Font(color='EEFFEE')
                header_background_fill = PatternFill(patternType='solid', fgColor='336633')


                # ヘッダー出力
                for column_letter in output_column_letters:
                    cell = ws[f'{column_letter}1']
                    cell.value = output_columns[column_letter]['label']
                    cell.font = header_font
                    cell.fill = header_background_fill
                    cell.alignment = output_columns[column_letter]['align']


                for sorted_row_no, (row_no, row) in enumerate(df.iterrows()):
                    #print(f"{row_th=} {type(row)=}")
                    row_th = sorted_row_no + 2

                    upper_limit_coins_letter = None
                    h_time = None
                    t_time = None
                    for column_letter in output_column_letters:

                        cell = ws[f'{column_letter}{row_th}']
                        name = output_columns[column_letter]['name']

                        # 元データにある列
                        if name is not None:
                            cell.value = row[output_columns[column_letter]['name']]

                        # 元データにない列
                        else:
                            label = output_columns[column_letter]['label']

                            if label == '必要表回数':
                                h_time = math.ceil(row['span'] / row['h_step'])
                                cell.value = h_time

                            elif label == '必要ｳﾗ回数':
                                t_time = math.ceil(row['span'] / row['t_step'])
                                cell.value = t_time

                            elif label == '最短対局数':
                                cell.value = SeriesRule.let_shortest_coins(h_step=row['h_step'], t_step=row['t_step'], span=row['span'], turn_system_id=spec.turn_system_id)

                            elif label == '対局数上限':
                                # あとで設定
                                upper_limit_coins_letter = column_letter

                            elif label == 'Ａさんの優勝率（失敗抜）':
                                cell.value = row['a_victory_rate_by_trio'] / (1 - row['no_victory_rate'])
                            
                            elif label == 'Ｂさんの優勝率（失敗抜）':
                                cell.value = row['b_victory_rate_by_trio'] / (1 - row['no_victory_rate'])


                        cell.alignment = output_columns[column_letter]['align']


                    # 対局数上限
                    cell = ws[f'{upper_limit_coins_letter}{row_th}']
                    cell.value = SeriesRule.let_upper_limit_coins_without_failure_rate(spec=spec, h_time=h_time, t_time=t_time)


                # ウィンドウ枠の固定
                ws.freeze_panes = 'A2'


                # データ部背景色
                # -------------
                light_red_fill = PatternFill(patternType='solid', fgColor='FFEEEE')
                light_green_fill = PatternFill(patternType='solid', fgColor='EEFFEE')
                light_blue_fill = PatternFill(patternType='solid', fgColor='EEEEFF')


                # データ部罫線出力
                # ---------------
                black_side = Side(style='thick', color='333333')
                left_border = Border(left=black_side)
                right_border = Border(right=black_side)
                top_border = Border(top=black_side)

                for row_th in range(2, ws.max_row + 1):

                    # ［表点］
                    cell = ws[f'A{row_th}']
                    cell.border = left_border
                    cell.fill = light_red_fill

                    # ［ｳﾗ点］
                    cell = ws[f'B{row_th}']
                    cell.fill = light_red_fill

                    # ［優勝点］
                    cell = ws[f'C{row_th}']
                    cell.border = right_border
                    cell.fill = light_red_fill

                    # ［対局数上限］
                    cell = ws[f'G{row_th}']
                    cell.fill = light_green_fill

                    # ［Ａさんの優勝率（失敗込）］
                    cell = ws[f'H{row_th}']
                    cell.border = left_border

                    # ［Ａさんの優勝率（失敗抜）］
                    cell = ws[f'K{row_th}']
                    cell.border = left_border
                    cell.fill = light_blue_fill

                    # ［Ｂさんの優勝率（失敗抜）］
                    cell = ws[f'L{row_th}']
                    cell.border = right_border
                    cell.fill = light_blue_fill


                # ワークブックの保存
                wb.save(detail_wb_file_path)
                print(f"[{datetime.datetime.now()}] please look `{detail_wb_file_path}`")


            except PermissionError as e:
                message = f"[{datetime.datetime.now()}] ファイルが他で開かれているのかも？ {detail_csv_file_path=} {detail_wb_file_path=} {e=}"
                print(message)
                # スタックトレース表示
                print(traceback.format_exc())

                log_file_path = VictoryRateDetailFilePaths.as_log(spec=spec)
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"{message}\n")    # ファイルへ出力

                # １分休む
                seconds = 60
                print(f"[{datetime.datetime.now()}] retry after {seconds} seconds")
                time.sleep(seconds)


            except Exception as e:
                message = f"[{datetime.datetime.now()}] 予期せぬ例外 {detail_csv_file_path=} {detail_wb_file_path=} {e=}"
                print(message)
                # スタックトレース表示
                print(traceback.format_exc())

                log_file_path = VictoryRateDetailFilePaths.as_log(spec=spec)
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"{message}\n")    # ファイルへ出力

                # １分休む
                seconds = 60
                print(f"[{datetime.datetime.now()}] retry after {seconds} seconds")
                time.sleep(seconds)


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
